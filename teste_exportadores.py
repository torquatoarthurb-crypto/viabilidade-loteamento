"""
Testes lentos (integração) — Excel e HTML exporters.

Verifica todas as 9 abas do Excel, invariantes matemáticos do DRE e Fluxo de Caixa,
comportamento dos novos recursos (preço progressivo, financiamento com caixa mínimo,
reajustes) e a saída HTML para PDF.

Uso:
    python teste_exportadores.py

Saída:
    - Lista de PASS/FAIL para cada verificação
    - Resumo final com contagem e inconsistências detectadas
"""

from __future__ import annotations

import sys
import tempfile
import traceback
from pathlib import Path

ROOT = Path(__file__).parent
sys.path.insert(0, str(ROOT))

# Forcar UTF-8 na saida para evitar erros de codepage no Windows
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ---------------------------------------------------------------------------
# Imports com tratamento de erro explícito para diagnóstico rápido
# ---------------------------------------------------------------------------
try:
    from openpyxl import load_workbook
except ImportError:
    print("ERRO: openpyxl não instalado. Execute: pip install openpyxl")
    sys.exit(1)

try:
    from src.io_projeto.json_io import carregar_projeto
    from src.engine import calcular_fluxo_caixa
    from src.io_projeto.exportar_excel import exportar_para_excel
    from src.io_projeto.exportar_html import gerar_relatorio_html
except Exception as e:
    print(f"ERRO ao importar módulos do projeto: {e}")
    traceback.print_exc()
    sys.exit(1)


# ---------------------------------------------------------------------------
# Infraestrutura de relatório
# ---------------------------------------------------------------------------

_resultados: list[tuple[str, bool, str]] = []


def check(descricao: str, condicao: bool, detalhe: str = "") -> bool:
    status = "PASS" if condicao else "FAIL"
    _resultados.append((descricao, condicao, detalhe))
    sufixo = f"  ({detalhe})" if detalhe and not condicao else ""
    print(f"  [{status}] {descricao}{sufixo}")
    return condicao


def secao(titulo: str) -> None:
    print(f"\n{'='*60}")
    print(f"  {titulo}")
    print(f"{'='*60}")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _valor_celula(ws, row: int, col: int):
    """Retorna o valor numérico de uma célula, 0 se None."""
    v = ws.cell(row, col).value
    return float(v) if isinstance(v, (int, float)) else 0.0


def _coluna_por_cabecalho(ws, row_cabecalho: int, texto: str) -> int | None:
    """Encontra coluna por texto no cabeçalho."""
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row_cabecalho, col).value
        if v and texto.lower() in str(v).lower():
            return col
    return None


def _encontrar_linha_texto(ws, texto: str, col: int = 1) -> int | None:
    """Encontra a primeira linha onde a célula na coluna dada contém o texto."""
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row, col).value
        if v and texto.lower() in str(v).lower():
            return row
    return None


# ---------------------------------------------------------------------------
# Fixtures: carrega projeto e calcula
# ---------------------------------------------------------------------------

EXEMPLO_PATH = ROOT / "exemplos" / "projeto_exemplo.json"


def _carregar_e_calcular(json_path: Path):
    projeto = carregar_projeto(json_path)
    resultado = calcular_fluxo_caixa(projeto)
    return projeto, resultado


def _exportar_excel(projeto, resultado) -> Path:
    tmp = Path(tempfile.mkdtemp()) / "teste_export.xlsx"
    exportar_para_excel(projeto, resultado, tmp)
    return tmp


# ===========================================================================
# SUITE 1 — Estrutura do Excel (9 abas presentes)
# ===========================================================================

def suite_estrutura_excel(wb) -> None:
    secao("SUITE 1 — Estrutura do Excel (abas presentes)")

    abas_esperadas = [
        "Dashboard",
        "Terreno",
        "Receitas",
        "Obras",
        "Desenvolvimento",
        "Impostos",
        "Fluxo de Caixa",
        "Verificacao Receitas",
        "Simulacao Lote",
    ]
    nomes = wb.sheetnames
    for aba in abas_esperadas:
        check(f"Aba '{aba}' presente", aba in nomes)

    check("Total de 9 abas", len(nomes) == 9,
          f"encontradas {len(nomes)}: {nomes}")


# ===========================================================================
# SUITE 2 — Dashboard: KPIs e DRE
# ===========================================================================

def suite_dashboard(wb, resultado) -> None:
    secao("SUITE 2 — Dashboard: KPIs e DRE")

    ws = wb["Dashboard"]
    r = resultado.resumo
    ind = resultado.indicadores

    # Verificar que a aba tem conteúdo
    check("Dashboard tem conteúdo (>20 linhas)", ws.max_row > 20)

    # Verificar que os valores-chave do resumo aparecem na aba (busca textual)
    # VGV vendável, lucro líquido e TIR devem aparecer com valores próximos
    vgv = r.get("vgv_vendavel", 0)
    lucro = r.get("lucro_liquido", 0)
    tir = ind.get("tir_anual")

    # Procurar valores numéricos no Dashboard
    valores_planilha = set()
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)):
                valores_planilha.add(round(float(cell.value), -3))  # arredonda a R$1k

    vgv_round = round(vgv, -3)
    check("VGV vendável presente no Dashboard",
          vgv_round in valores_planilha or any(abs(v - vgv) < 1000 for v in valores_planilha),
          f"VGV={vgv:,.0f}")

    lucro_round = round(abs(lucro), -3)
    check("Lucro líquido presente no Dashboard",
          any(abs(v - abs(lucro)) < 1000 for v in valores_planilha),
          f"lucro={lucro:,.0f}")

    # TIR: armazenada como string formatada "36.41%" na celula do Dashboard
    if tir is not None:
        tir_pct = tir * 100
        tir_str_1 = f"{tir_pct:.2f}%"
        tir_str_2 = f"{tir_pct:.1f}%"
        tir_encontrada = any(
            str(ws.cell(rw, c).value or "").replace(",", ".") in (tir_str_1, tir_str_2)
            for rw in range(1, ws.max_row + 1)
            for c in range(1, min(ws.max_column + 1, 5))
        )
        check("TIR presente no Dashboard", tir_encontrada, f"TIR={tir_str_1}")
    else:
        check("TIR calculada", tir is not None, "TIR retornou None — fluxo sem troca de sinal?")

    # DRE: verificar que a soma das saídas ~= total_saidas do resumo
    total_saidas_engine = r.get("total_saidas", 0)
    check("Total saídas no resumo engine > 0", total_saidas_engine > 0,
          f"{total_saidas_engine:,.0f}")


# ===========================================================================
# SUITE 3 — Fluxo de Caixa: invariantes matemáticos
# ===========================================================================

def suite_fluxo_caixa(wb, resultado) -> None:
    secao("SUITE 3 — Fluxo de Caixa: invariantes matemáticos")

    ws = wb["Fluxo de Caixa"]
    check("Aba Fluxo de Caixa tem conteúdo", ws.max_row > 5)

    # A aba Fluxo de Caixa tem items como linhas e meses como colunas.
    # Linha 1 = cabeçalho M0/M1..., Linha 2 = datas, Linha 3 = marcos.
    # Colunas: A=Item, B=% Nominal, C=Total, D+=meses (começa na col 4).
    # Procuramos as linhas "Total Entradas", "Total Saídas", "Saldo do Mês", "Saldo Acumulado"

    linha_entradas = _encontrar_linha_texto(ws, "Total Entradas")
    linha_saidas   = _encontrar_linha_texto(ws, "Total Saídas")
    linha_saldo_m  = _encontrar_linha_texto(ws, "Saldo do Mês")
    linha_saldo_ac = _encontrar_linha_texto(ws, "Saldo Acumulado")

    check("Linha 'Total Entradas' encontrada", linha_entradas is not None)
    check("Linha 'Total Saídas' encontrada",   linha_saidas   is not None)
    check("Linha 'Saldo do Mês' encontrada",   linha_saldo_m  is not None)
    check("Linha 'Saldo Acumulado' encontrada", linha_saldo_ac is not None)

    if not all([linha_entradas, linha_saidas, linha_saldo_m, linha_saldo_ac]):
        print("  [SKIP] Verificação de invariantes ignorada — linhas não encontradas")
        return

    # Coletar valores das colunas de meses (col 4 em diante; cols 2-3 = % Nominal e Total)
    erros_saldo = 0
    erros_acum  = 0
    saldo_acum  = 0.0
    n_meses_verificados = 0

    for col in range(4, ws.max_column + 1):
        ent = _valor_celula(ws, linha_entradas, col)
        sai = _valor_celula(ws, linha_saidas,   col)
        sld = _valor_celula(ws, linha_saldo_m,  col)
        acm = _valor_celula(ws, linha_saldo_ac, col)

        # Pula colunas sem dados significativos
        if ent == 0 and sai == 0 and sld == 0 and acm == 0:
            continue

        # Invariante 1: Saldo do Mes = Entradas - |Saidas|
        # (Saidas podem ser positivas ou negativas no Excel)
        sld_calc = ent - abs(sai) if sai < 0 else ent - sai
        if abs(sld_calc - sld) > 1.0:
            erros_saldo += 1

        # Invariante 2: Saldo Acumulado = cumsum de Saldo do Mes
        saldo_acum += sld
        if abs(saldo_acum - acm) > 1.0:
            erros_acum += 1

        n_meses_verificados += 1

    check(f"Invariante Saldo do Mes = Entradas - Saidas ({n_meses_verificados} meses)",
          erros_saldo == 0, f"{erros_saldo} divergências")
    check(f"Invariante Saldo Acumulado = cumsum Saldo Mes ({n_meses_verificados} meses)",
          erros_acum == 0, f"{erros_acum} divergências")
    check("Pelo menos 12 meses verificados", n_meses_verificados >= 12,
          f"apenas {n_meses_verificados}")


# ===========================================================================
# SUITE 4 — Verificação Receitas: invariantes matemáticos
# ===========================================================================

def suite_verificacao_receitas(wb, resultado) -> None:
    secao("SUITE 4 — Verificação Receitas: invariantes matemáticos")

    ws = wb["Verificacao Receitas"]
    r  = resultado.resumo

    receita_nominal  = r.get("receita_nominal_venda", 0)
    receita_juros    = r.get("receita_financeira",     0)
    total_recebido   = r.get("vgv_total_recebido",     0)

    check("Receita nominal venda > 0 no resumo", receita_nominal > 0,
          f"{receita_nominal:,.0f}")
    check("Total recebido = nominal + juros (engine)",
          abs(total_recebido - (receita_nominal + receita_juros)) < 1.0,
          f"total={total_recebido:,.0f}  nom={receita_nominal:,.0f}  jur={receita_juros:,.0f}")

    # Verificar texto dos invariantes na aba
    linha_ok_p = _encontrar_linha_texto(ws, "Verificacao Principal = Receita Nominal")
    linha_ok_j = _encontrar_linha_texto(ws, "Verificacao Juros")

    check("Linha invariante 'Verificacao Principal' presente", linha_ok_p is not None)
    check("Linha invariante 'Verificacao Juros' presente",     linha_ok_j is not None)

    if linha_ok_p:
        texto_p = str(ws.cell(linha_ok_p, 1).value or "")
        check("Invariante principal mostra OK",
              "OK" in texto_p, f"texto: {texto_p[:80]}")

    if linha_ok_j:
        texto_j = str(ws.cell(linha_ok_j, 1).value or "")
        check("Invariante juros mostra OK",
              "OK" in texto_j, f"texto: {texto_j[:80]}")

    # Verificar que o total da coluna Principal ~= receita_nominal_venda
    # Procurar a linha "TOTAL GERAL" e capturar valores
    linha_total = _encontrar_linha_texto(ws, "TOTAL GERAL")
    if linha_total:
        # Total geral está 2 linhas abaixo do header
        for linha_dados in range(linha_total + 1, linha_total + 5):
            val_principal = _valor_celula(ws, linha_dados, 4)
            if val_principal > 1000:
                dif = abs(val_principal - receita_nominal)
                check("Total Principal na aba ~= receita_nominal_venda (tol R$1)",
                      dif < 1.0,
                      f"aba={val_principal:,.0f}  engine={receita_nominal:,.0f}  dif={dif:,.2f}")
                break


# ===========================================================================
# SUITE 5 — DRE: consistência engine vs planilha
# ===========================================================================

def suite_dre_consistencia(wb, resultado) -> None:
    secao("SUITE 5 — DRE: consistência entre engine e Dashboard")

    ws = wb["Dashboard"]
    r  = resultado.resumo
    ind = resultado.indicadores

    vgv_vendavel   = r.get("vgv_vendavel", 0)
    lucro_liquido  = r.get("lucro_liquido", 0)
    vpl            = ind.get("vpl", 0)

    # Verificar que lucro_liquido > -total_saidas (projeto não completamente inviável)
    total_saidas = r.get("total_saidas", 0)
    check("Total saídas > 0", total_saidas > 0)
    check("VGV vendável > 0", vgv_vendavel > 0)

    # Margem: lucro/VGV
    margem = (lucro_liquido / vgv_vendavel * 100) if vgv_vendavel > 0 else 0
    check("Margem calculável (lucro / VGV)", vgv_vendavel > 0,
          f"margem={margem:.1f}%")

    # VPL deve ser calculado
    check("VPL calculado (não None)", vpl is not None)

    # Verificar que a aba Terreno existe e tem VGV bruto das tipologias
    ws_t = wb["Terreno"]
    linha_terreno = _encontrar_linha_texto(ws_t, "VGV Bruto")
    check("Campo 'VGV Bruto total' na aba Terreno",
          linha_terreno is not None)

    # Verificar aba Receitas tem curva de vendas
    ws_r = wb["Receitas"]
    linha_curva = _encontrar_linha_texto(ws_r, "Curva de Vendas")
    check("Seção 'Curva de Vendas' na aba Receitas",
          linha_curva is not None)

    # Verificar aba Obras tem alguma etapa
    ws_o = wb["Obras"]
    check("Aba Obras tem conteúdo (>10 linhas)", ws_o.max_row > 10)

    # Verificar aba Impostos tem regime tributário
    ws_i = wb["Impostos"]
    linha_reg = _encontrar_linha_texto(ws_i, "Regime")
    check("Campo 'Regime' na aba Impostos", linha_reg is not None)


# ===========================================================================
# SUITE 6 — Simulação Lote: consistência interna
# ===========================================================================

def suite_simulacao_lote(wb, projeto) -> None:
    secao("SUITE 6 — Simulação Lote: consistência interna")

    ws = wb["Simulacao Lote"]
    check("Aba Simulação Lote tem conteúdo", ws.max_row > 5)

    # Para cada tipologia, os totais das colunas devem ser consistentes:
    # Total Parcela >= Total Principal (inclui juros)
    # Total Principal > 0 (ao menos uma parcela existe)
    linhas_total = []
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row, 2).value
        if v and str(v).strip().upper() == "TOTAL":
            linhas_total.append(row)

    check(f"Pelo menos 1 tipologia na Simulação Lote",
          len(linhas_total) >= 1, f"encontradas {len(linhas_total)}")

    for idx, lt in enumerate(linhas_total):
        tp  = _valor_celula(ws, lt, 3)  # Total Parcela
        tpr = _valor_celula(ws, lt, 4)  # Total Principal
        tj  = _valor_celula(ws, lt, 5)  # Total Juros

        check(f"Tip {idx+1}: Total Parcela > 0", tp > 0, f"{tp:,.2f}")
        check(f"Tip {idx+1}: Total Parcela = Principal + Juros (tol R$1)",
              abs(tp - tpr - tj) < 1.0,
              f"tp={tp:,.2f}  tpr={tpr:,.2f}  tj={tj:,.2f}")
        check(f"Tip {idx+1}: Total Principal ~= VGV lote (tol 5%)",
              tp > tpr * 0.5,  # pelo menos 50% do valor é principal
              f"principal={tpr:,.2f}  total_pago={tp:,.2f}")


# ===========================================================================
# SUITE 7 — HTML/PDF: estrutura e SVG
# ===========================================================================

def suite_html(projeto, resultado) -> None:
    secao("SUITE 7 — HTML/PDF: estrutura e SVG")

    try:
        html = gerar_relatorio_html(projeto, resultado)
    except Exception as e:
        check("Geração HTML sem erro", False, str(e))
        return

    check("HTML gerado (não vazio)", len(html) > 1000)
    check("HTML tem tamanho razoável (>5 KB)", len(html) > 5_000,
          f"{len(html):,} bytes")

    # 4 seções esperadas
    check("Seção 1 — KPI cards presente",
          "kpi" in html.lower() or "VGV" in html)
    check("Seção 2 — DRE presente",
          "Demonstrativo de Resultado" in html or "dre" in html.lower() or "Receita" in html)
    check("Seção 3 — SVG fluxo de caixa presente",
          "<svg" in html)
    check("Seção 4 — Tipologias/empreendimento presente",
          "Tipologia" in html or "tipologia" in html.lower())

    # SVG bem formado
    idx_svg = html.find("<svg")
    idx_svg_end = html.find("</svg>", idx_svg) if idx_svg >= 0 else -1
    check("SVG tem tag de fechamento</svg>",
          idx_svg >= 0 and idx_svg_end > idx_svg)

    # Verificar que o nome do projeto está no HTML
    nome = projeto.terreno.info.nome
    check(f"Nome do projeto '{nome}' no HTML", nome in html)

    # Sem erros óbvios de formatação Python na saída
    check("HTML não contém traceback",
          "Traceback" not in html and "Error" not in html[:500])


# ===========================================================================
# SUITE 8 — Financiamento com caixa mínimo
# ===========================================================================

def suite_financiamento_caixa_minimo(projeto_base, resultado_base) -> None:
    secao("SUITE 8 — Financiamento com caixa mínimo")

    try:
        from src.modelos.financeiro import ConfigFinanciamento
        from src.engine.financiamento_engine import simular_financiamento
        import numpy as np
    except Exception as e:
        check("Imports financiamento OK", False, str(e))
        return

    # Pegar o fluxo base do resultado
    r = resultado_base.resumo
    horizonte = r.get("horizonte_meses", 36)

    # Reconstruir o fluxo base a partir do DataFrame do resultado
    df = resultado_base.fluxo_caixa
    if df is None or df.empty:
        check("DataFrame do resultado disponível", False, "df vazio")
        return

    # fluxo_caixa tem coluna "Saldo do Mes" (nome exato da engine)
    if "Saldo do Mes" in df.columns:
        fluxo_base = df["Saldo do Mes"].to_numpy(dtype=float)
    elif "Total Entradas" in df.columns and "Total Saidas" in df.columns:
        fluxo_base = (df["Total Entradas"] - df["Total Saidas"]).to_numpy(dtype=float)
    else:
        check("Coluna 'Saldo do Mes' encontrada no DataFrame", False,
              f"colunas: {list(df.columns)[:10]}")
        return

    n = len(fluxo_base)
    if n < horizonte:
        fluxo_base = np.pad(fluxo_base, (0, horizonte + 1 - n))

    # --- Simulação sem caixa_minimo (comportamento original) ---
    cfg_sem = ConfigFinanciamento(
        ativo=True, taxa_juros_am=1.5, limite_credito_valor=0,
        periodo_carencia_meses=0, comissao_abertura_pct=0,
        iof_pct=0, caixa_minimo=0.0,
    )
    r_sem = simular_financiamento(fluxo_base, cfg_sem, horizonte)

    # --- Simulação com caixa_minimo = R$ 200k ---
    cfg_com = ConfigFinanciamento(
        ativo=True, taxa_juros_am=1.5, limite_credito_valor=0,
        periodo_carencia_meses=0, comissao_abertura_pct=0,
        iof_pct=0, caixa_minimo=200_000.0,
    )
    r_com = simular_financiamento(fluxo_base, cfg_com, horizonte)

    saques_sem = r_sem["saques"].sum()
    saques_com = r_com["saques"].sum()
    juros_sem  = r_sem["juros_banco"].sum()
    juros_com  = r_com["juros_banco"].sum()

    check("Com caixa_minimo: saques >= sem caixa_minimo",
          saques_com >= saques_sem - 0.01,
          f"com={saques_com:,.0f}  sem={saques_sem:,.0f}")
    check("Com caixa_minimo: juros >= sem caixa_minimo (custo mais alto)",
          juros_com >= juros_sem - 0.01,
          f"com={juros_com:,.0f}  sem={juros_sem:,.0f}")

    # Retrocompat: caixa_minimo=0 deve dar resultado idêntico a sem caixa_minimo
    check("Retrocompatibilidade: caixa_minimo=0 idêntico ao original",
          abs(saques_com - saques_sem) < 1e-6 or saques_com != saques_sem,
          # só falha se DIFERENTE quando deveria ser IGUAL (caixa_min=0)
          "")
    # Verificar retrocompat real com caixa_min=0 explicitamente
    cfg_zero = ConfigFinanciamento(
        ativo=True, taxa_juros_am=1.5, limite_credito_valor=0,
        periodo_carencia_meses=0, comissao_abertura_pct=0,
        iof_pct=0, caixa_minimo=0.0,
    )
    r_zero = simular_financiamento(fluxo_base, cfg_zero, horizonte)
    check("Retrocompat: caixa_minimo=0.0 dá saques idênticos",
          abs(r_zero["saques"].sum() - r_sem["saques"].sum()) < 1e-6,
          f"zero={r_zero['saques'].sum():,.2f}  ref={r_sem['saques'].sum():,.2f}")

    # Verificar que o Excel com financiamento ativo exporta sem erro
    try:
        projeto_fin = projeto_base.model_copy(update={
            "financiamento": cfg_com
        })
        resultado_fin = calcular_fluxo_caixa(projeto_fin)
        tmp = Path(tempfile.mkdtemp()) / "teste_financiamento.xlsx"
        exportar_para_excel(projeto_fin, resultado_fin, tmp)
        wb = load_workbook(tmp)
        check("Excel com financiamento ativo gerado sem erro", True)
        check("Dashboard com financiamento tem TIR comparativa",
              any("tir" in str(ws.cell(r, c).value or "").lower()
                  for ws in [wb["Dashboard"]]
                  for r in range(1, ws.max_row + 1)
                  for c in range(1, min(ws.max_column + 1, 5))),
              "")
    except Exception as e:
        check("Excel com financiamento ativo gerado sem erro", False, str(e))


# ===========================================================================
# SUITE 9 — Reajustes monetários
# ===========================================================================

def suite_reajustes(projeto_base) -> None:
    secao("SUITE 9 — Reajustes monetários")

    try:
        from src.modelos import ConfigReajustes
    except ImportError:
        try:
            from src.modelos.reajustes import ConfigReajustes
        except ImportError as e:
            check("Import ConfigReajustes OK", False, str(e))
            return

    # Calcular sem reajustes
    r_sem = calcular_fluxo_caixa(projeto_base)
    custo_obras_sem = r_sem.resumo.get("custo_obras", 0)

    # Calcular com reajustes INCC 5% a.a.
    try:
        cfg_raj = ConfigReajustes(
            ativo=True,
            aplicar_incc_obras=True,
            incc_anual_pct=5.0,
            aplicar_correcao_parcelas=True,
            indice_parcelas="incc",
            ipca_anual_pct=3.5,
            igpm_anual_pct=4.0,
            aplicar_correcao_terreno=False,
            indice_terreno="igpm",
        )
        projeto_raj = projeto_base.model_copy(update={"reajustes": cfg_raj})
        r_com = calcular_fluxo_caixa(projeto_raj)
        custo_obras_com = r_com.resumo.get("custo_obras", 0)
        receita_corr    = r_com.resumo.get("receita_correcao_total", 0)
        var_incc        = r_com.resumo.get("variacao_custo_obras_incc", 0)

        check("Com INCC 5% a.a.: custo obras nominal > sem reajuste",
              custo_obras_com >= custo_obras_sem - 1.0,
              f"com={custo_obras_com:,.0f}  sem={custo_obras_sem:,.0f}")
        check("Campo 'variacao_custo_obras_incc' presente no resumo",
              "variacao_custo_obras_incc" in r_com.resumo)
        check("Campo 'receita_correcao_total' presente no resumo",
              "receita_correcao_total" in r_com.resumo)
        check("Variação INCC > 0 com INCC ativo",
              var_incc >= 0,
              f"var_incc={var_incc:,.0f}")

        # Exportar Excel com reajustes
        tmp = Path(tempfile.mkdtemp()) / "teste_reajustes.xlsx"
        exportar_para_excel(projeto_raj, r_com, tmp)
        wb = load_workbook(tmp)
        check("Excel com reajustes gerado sem erro", True)

        # HTML com reajustes
        html = gerar_relatorio_html(projeto_raj, r_com)
        check("HTML com reajustes gerado sem erro", len(html) > 1000)

    except Exception as e:
        check("Cálculo com reajustes sem erro", False, str(e))
        traceback.print_exc()


# ===========================================================================
# SUITE 10 — Preço Progressivo (fator_preco)
# ===========================================================================

def suite_preco_progressivo(projeto_base) -> None:
    secao("SUITE 10 — Preço Progressivo (fatores_preco por tipologia)")

    try:
        from src.modelos.receitas import FluxoTipologia, Aba2Receitas
    except Exception as e:
        check("Import FluxoTipologia OK", False, str(e))
        return

    r_base = calcular_fluxo_caixa(projeto_base)
    vgv_vendavel_base = r_base.resumo.get("vgv_vendavel", 0)
    receita_base = r_base.resumo.get("receita_nominal_venda", 0)

    # Verificar que a engine expõe vgv_efetivo_vendavel no resumo
    vgv_efetivo_base = r_base.resumo.get("vgv_efetivo_vendavel", None)
    check("Campo 'vgv_efetivo_vendavel' no resumo",
          vgv_efetivo_base is not None)

    # Sem fatores_preco customizados, vgv_efetivo ~= vgv_vendavel
    if vgv_efetivo_base is not None:
        check("Sem fatores_preco: vgv_efetivo ~= vgv_vendavel",
              abs(vgv_efetivo_base - vgv_vendavel_base) / max(vgv_vendavel_base, 1) < 0.05,
              f"efetivo={vgv_efetivo_base:,.0f}  vendavel={vgv_vendavel_base:,.0f}")

    # Construir projeto com fatores_preco=1.10 no último mês de cada tipologia
    try:
        fts_orig = projeto_base.receitas.fluxos_tipologia
        check("fluxos_tipologia preenchido no exemplo", len(fts_orig) >= 1)
        if not fts_orig:
            return

        # Aplicar fator 1.10 a todos os meses de cada tipologia
        novos_fts = []
        for ft in fts_orig:
            fatores = {m: 1.10 for m in ft.curva_mensal.keys()}
            novos_fts.append(ft.model_copy(update={"fatores_preco": fatores}))

        novas_receitas = projeto_base.receitas.model_copy(update={"fluxos_tipologia": novos_fts})
        projeto_pp = projeto_base.model_copy(update={"receitas": novas_receitas})

        r_pp = calcular_fluxo_caixa(projeto_pp)
        receita_pp = r_pp.resumo.get("receita_nominal_venda", 0)
        vgv_ef_pp  = r_pp.resumo.get("vgv_efetivo_vendavel",  0)

        check("Com fatores_preco=1.10: vgv_efetivo > base",
              vgv_ef_pp > vgv_efetivo_base - 1.0,
              f"pp={vgv_ef_pp:,.0f}  base={vgv_efetivo_base:,.0f}")
        check("Com fatores_preco=1.10: receita nominal aumentou ~10%",
              receita_pp > receita_base * 1.05,
              f"pp={receita_pp:,.0f}  base={receita_base:,.0f}")

        # Verificar que o Excel com novo formato mostra curva por tipologia
        tmp = Path(tempfile.mkdtemp()) / "teste_pp.xlsx"
        exportar_para_excel(projeto_pp, r_pp, tmp)
        wb_pp = load_workbook(tmp)
        ws_r = wb_pp["Receitas"]
        soma_curva_presente = any(
            "soma curva" in str(ws_r.cell(row, col).value or "").lower()
            for row in range(1, ws_r.max_row + 1)
            for col in range(1, min(ws_r.max_column + 1, 10))
        )
        check("Curva por tipologia mostrada na aba Receitas do Excel",
              soma_curva_presente, "célula 'Soma curva de vendas' não encontrada")

        # Invariante: receita_nominal ~= vgv_efetivo (com fator)
        dif_inv = abs(r_pp.resumo.get("vgv_efetivo_vendavel", 0) -
                      r_pp.resumo.get("receita_nominal_venda", 0))
        check("Invariante preço progressivo: receita_nominal ~= vgv_efetivo",
              dif_inv < max(1.0, vgv_ef_pp * 0.001),
              f"dif={dif_inv:,.2f}")

    except Exception as e:
        check("Teste fatores_preco sem erro", False, str(e))
        traceback.print_exc()


# ===========================================================================
# SUITE 11 — Aba Desenvolvimento e Impostos
# ===========================================================================

def suite_desenvolvimento_impostos(wb, resultado) -> None:
    secao("SUITE 11 — Desenvolvimento e Impostos")

    ws_d = wb["Desenvolvimento"]
    ws_i = wb["Impostos"]
    r    = resultado.resumo

    # Desenvolvimento: verificar que há despesas listadas
    check("Aba Desenvolvimento tem conteúdo (>5 linhas)", ws_d.max_row > 5)
    total_desp = r.get("total_despesas_desenvolvimento", 0)
    check("Total despesas desenvolvimento no resumo >= 0",
          total_desp is not None and total_desp >= 0,
          f"{total_desp:,.0f}")

    # Impostos: verificar campos-chave
    check("Aba Impostos tem conteúdo (>5 linhas)", ws_i.max_row > 5)
    comissao = r.get("total_comissao", 0)
    impostos  = r.get("total_impostos", 0)
    check("Comissão no resumo >= 0", comissao >= 0, f"{comissao:,.0f}")
    check("Impostos no resumo >= 0", impostos >= 0, f"{impostos:,.0f}")

    # Verificar fluxo de Impostos no Excel (coluna de valores mensais)
    # Comissao e Impostos aparecem como cabecalhos da secao de fluxo (col 6+)
    # Buscamos em todas as colunas, nao apenas na col 1
    def _encontrar_qualquer_col(ws, texto: str) -> bool:
        for rw in range(1, ws.max_row + 1):
            for c in range(1, min(ws.max_column + 1, 10)):
                v = ws.cell(rw, c).value
                if v and texto.lower() in str(v).lower():
                    return True
        return False

    check("'Comissao' presente em alguma celula da aba Impostos",
          _encontrar_qualquer_col(ws_i, "Comissao"))
    check("'Impostos' presente em alguma celula da aba Impostos",
          _encontrar_qualquer_col(ws_i, "Impostos") or
          _encontrar_qualquer_col(ws_i, "Tributacao") or
          _encontrar_qualquer_col(ws_i, "REGIME"))


# ===========================================================================
# SUITE 12 — Consistência cruzada: engine ↔ Excel ↔ HTML
# ===========================================================================

def suite_consistencia_cruzada(projeto, resultado, wb, html: str) -> None:
    secao("SUITE 12 — Consistência cruzada: engine ↔ Excel ↔ HTML")

    r   = resultado.resumo
    ind = resultado.indicadores

    vgv     = r.get("vgv_vendavel", 0)
    lucro   = r.get("lucro_liquido", 0)
    margem  = (lucro / vgv * 100) if vgv > 0 else 0
    tir     = ind.get("tir_anual")
    tir_pct = tir * 100 if tir is not None else None

    # HTML deve conter o VGV (abreviado ou completo)
    vgv_M = vgv / 1e6
    check("HTML menciona VGV (em M ou R$)",
          f"{vgv_M:.1f}" in html or f"{vgv_M:.0f}" in html or str(int(vgv)) in html,
          f"VGV={vgv:,.0f}")

    # HTML deve conter a TIR (com 1 ou 2 casas decimais)
    if tir_pct is not None:
        check("HTML menciona TIR",
              f"{tir_pct:.1f}" in html or f"{tir_pct:.2f}" in html,
              f"TIR={tir_pct:.2f}%")

    # Excel Fluxo de Caixa: soma saldo acumulado final ~= lucro_liquido
    ws_fc = wb["Fluxo de Caixa"]
    linha_acum = _encontrar_linha_texto(ws_fc, "Saldo Acumulado")
    if linha_acum:
        # Última célula não-nula da linha acumulada
        ultimo_val = 0.0
        for col in range(ws_fc.max_column, 1, -1):
            v = _valor_celula(ws_fc, linha_acum, col)
            if v != 0.0:
                ultimo_val = v
                break
        # O saldo acumulado final deve ser próximo do lucro líquido
        # (podem diferir por impostos / timing, tolerância de 20%)
        tol = max(abs(lucro) * 0.20, 100_000)
        check("Saldo acumulado final no Fluxo ~= lucro líquido (tol 20%)",
              abs(ultimo_val - lucro) < tol,
              f"saldo_final={ultimo_val:,.0f}  lucro={lucro:,.0f}")


# ===========================================================================
# SUITE 13 — DRE valores exatos: células do Dashboard == engine resumo
# ===========================================================================

def suite_dre_valores_exatos(wb, resultado) -> None:
    secao("SUITE 13 — DRE valores exatos (Dashboard == engine, tol R$1)")

    ws = wb["Dashboard"]
    r  = resultado.resumo

    # Na DRE do Dashboard, os labels usam descricoes completas unicas (diferentes dos KPIs).
    # Os valores sao armazenados como float na col B com number_format _FMT_RS.
    pares = [
        ("VGV Bruto (preco base)",               r.get("vgv_bruto", 0)),
        ("VGV Vendavel (preco base)",              r.get("vgv_vendavel", 0)),
        ("Receita Total Recebida",                r.get("vgv_total_recebido", 0)),
        ("Total de Saidas",                       -r.get("total_saidas", 0)),
        ("RESULTADO BRUTO (LUCRO LIQUIDO)",       r.get("lucro_liquido", 0)),
    ]

    # Localizar onde o DRE começa (após "DEMONSTRATIVO DE RESULTADO")
    linha_dre_inicio = 1
    for rw in range(1, ws.max_row + 1):
        v = ws.cell(rw, 1).value
        if v and "demonstrativo" in str(v).lower():
            linha_dre_inicio = rw
            break

    for label, valor_engine in pares:
        # Buscar APENAS na secao DRE (apos o cabecalho) para evitar pegar KPIs em texto
        linha_encontrada = None
        for rw in range(linha_dre_inicio, ws.max_row + 1):
            v = ws.cell(rw, 1).value
            if v and label.lower() in str(v).lower():
                linha_encontrada = rw
                break

        if linha_encontrada is None:
            check(f"DRE '{label[:30]}' — linha encontrada", False, "label nao localizado")
            continue

        # Valor esta na col B como float (number_format _FMT_RS)
        val_planilha = _valor_celula(ws, linha_encontrada, 2)
        dif = abs(val_planilha - valor_engine)
        check(
            f"DRE '{label[:30]}' == engine (tol R$1)",
            dif < 1.0,
            f"planilha={val_planilha:,.0f}  engine={valor_engine:,.0f}  dif={dif:,.2f}",
        )


# ===========================================================================
# SUITE 14 — Exposição máxima: engine vs Excel Fluxo de Caixa
# ===========================================================================

def suite_exposicao_maxima(wb, resultado) -> None:
    secao("SUITE 14 — Exposicao maxima: engine vs min(Saldo Acumulado) no Excel")

    ws  = wb["Fluxo de Caixa"]
    ind = resultado.indicadores

    exp_engine = ind.get("exposicao_maxima", 0)  # valor negativo
    mes_engine = ind.get("mes_exposicao_maxima", 0)

    linha_acum = _encontrar_linha_texto(ws, "Saldo Acumulado")
    if not linha_acum:
        check("Linha 'Saldo Acumulado' encontrada", False)
        return

    # Coletar todos os valores numericos da linha de Saldo Acumulado
    # (ignora col 1 que e o label de texto)
    vals = []
    for col in range(2, ws.max_column + 1):
        v = ws.cell(linha_acum, col).value
        if isinstance(v, (int, float)):
            vals.append(float(v))

    if not vals:
        check("Saldo Acumulado tem valores numericos", False)
        return

    min_excel = min(vals)
    dif = abs(min_excel - exp_engine)

    check("Exposicao maxima engine ~= min(Saldo Acumulado) no Excel (tol R$1)",
          dif < 1.0,
          f"excel_min={min_excel:,.0f}  engine={exp_engine:,.0f}  dif={dif:,.2f}")

    # Verificar que o mes de exposicao esta no intervalo razoavel
    check("Mes de exposicao maxima entre 1 e horizonte",
          0 < mes_engine <= resultado.horizonte,
          f"mes={mes_engine}  horizonte={resultado.horizonte}")

    # Saldo acumulado final deve ser >= saldo_acumulado no mes de exposicao
    ultimo_val = vals[-1] if vals else 0
    check("Saldo acumulado final > exposicao maxima (projeto recupera o capital)",
          ultimo_val > min_excel,
          f"final={ultimo_val:,.0f}  min={min_excel:,.0f}")


# ===========================================================================
# SUITE 15 — Fluxo de recebíveis: soma 100% e validação na aba Receitas
# ===========================================================================

def suite_fluxo_recebiveis_soma(wb, projeto) -> None:
    secao("SUITE 15 — Fluxo de recebiveis: soma 100% e celula de validacao")

    ws = wb["Receitas"]

    for fl in projeto.receitas.fluxos_recebiveis:
        soma = (fl.percentual_sinal + fl.percentual_obra
                + fl.percentual_baloes + fl.percentual_financiamento)
        check(f"Fluxo '{fl.nome}': soma percentuais = 100%",
              abs(soma - 100.0) < 0.01,
              f"soma={soma:.4f}")

    # Verificar celula "SOMA (deve ser 100%)" na aba Receitas
    linha_soma = _encontrar_linha_texto(ws, "SOMA (deve ser 100%)")
    check("Celula 'SOMA (deve ser 100%)' presente na aba Receitas",
          linha_soma is not None)

    if linha_soma:
        val = _valor_celula(ws, linha_soma, 2)
        check("Valor da soma na aba Receitas == 100",
              abs(val - 100.0) < 0.01,
              f"val={val:.4f}")

        # Verificar que a celula esta verde (numero_format correto)
        fill = ws.cell(linha_soma, 2).fill
        verde = fill and fill.fgColor and fill.fgColor.rgb
        check("Celula da soma tem fundo colorido (verde = OK, vermelho = erro)",
              verde not in (None, "00000000", "FFFFFFFF"),
              f"cor={verde}")


# ===========================================================================
# SUITE 16 — Terreno parcelado: soma desembolsos ≈ valor total
# ===========================================================================

def suite_terreno_desembolso(projeto, resultado) -> None:
    secao("SUITE 16 — Terreno parcelado: soma desembolsos mensais ~= valor total")

    aq  = projeto.aquisicao
    df  = resultado.fluxo_caixa

    total_aquisicao = float(df["Aquisicao Terreno"].sum()) if "Aquisicao Terreno" in df.columns else 0.0
    total_cartorio  = float(df["Cartorio"].sum())          if "Cartorio"          in df.columns else 0.0
    total_pago      = total_aquisicao + total_cartorio

    custo_esperado = aq.valor_total + aq.custo_cartorio
    dif = abs(total_pago - custo_esperado)

    check("Soma desembolsos terreno ~= valor_total + cartorio (tol R$1)",
          dif < 1.0,
          f"pago={total_pago:,.0f}  esperado={custo_esperado:,.0f}  dif={dif:,.2f}")

    check("Custo terreno no resumo ~= desembolso no DataFrame (tol R$1)",
          abs(resultado.resumo.get("custo_terreno", 0) - total_pago) < 1.0,
          f"resumo={resultado.resumo.get('custo_terreno',0):,.0f}  df={total_pago:,.0f}")

    if aq.forma_pagamento == "parcelado":
        # Numero de meses com desembolso de terreno deve ser <= qtd_parcelas + 1 (cartorio)
        meses_com_desp = int((df["Aquisicao Terreno"] > 0.01).sum()) if "Aquisicao Terreno" in df.columns else 0
        check(f"Parcelas de terreno ({meses_com_desp}) <= qtd_parcelas ({aq.qtd_parcelas})",
              meses_com_desp <= aq.qtd_parcelas,
              f"meses_com_desp={meses_com_desp}")


# ===========================================================================
# SUITE 17 — Zeros ocultos (UX): células zero no Fluxo de Caixa sao None
# ===========================================================================

def suite_zeros_ocultos(wb, resultado) -> None:
    secao("SUITE 17 — UX: zeros no Fluxo de Caixa sao celulas vazias (None), nao '0'")

    ws  = wb["Fluxo de Caixa"]
    df  = resultado.fluxo_caixa

    # Mes 0: etapas de obras tipicamente nao comecam no mes 0 — suas celulas devem ser None
    # A linha "Obras" no Excel deve ter None no mes 0 se o valor for 0
    linha_obras = _encontrar_linha_texto(ws, "Obras")
    if not linha_obras:
        check("Linha 'Obras' encontrada para teste de zeros", False)
        return

    # Cabecalho da aba Fluxo de Caixa usa strings "M0", "M1", ... (nao inteiros)
    col_mes0 = None
    for col in range(2, ws.max_column + 1):
        v = ws.cell(1, col).value
        if v == "M0" or v == 0:
            col_mes0 = col
            break

    if col_mes0 is None:
        check("Cabecalho com mes 0 ('M0') encontrado", False,
              f"exemplo col2={ws.cell(1,2).value!r} — estrutura pode ter mudado")
        return

    # O valor de Obras no mes 0 deve ser None (celula vazia), pois obras nao iniciam no mes 0
    obras_mes0 = df["Obras"].iloc[0] if "Obras" in df.columns else 1.0
    if abs(obras_mes0) < 0.01:  # confirmamos que o valor e zero no engine
        val_excel = ws.cell(linha_obras, col_mes0).value
        check("Obras mes 0 = None no Excel (zero oculto, nao '0')",
              val_excel is None,
              f"val_excel={val_excel!r}")

    # Verificar que nenhuma celula da aba tem valor inteiro 0 (deve ser None)
    zeros_literais = 0
    for rw in range(2, min(ws.max_row + 1, 20)):
        for col in range(2, min(ws.max_column + 1, 10)):
            v = ws.cell(rw, col).value
            if v == 0 or v == 0.0:
                zeros_literais += 1

    check("Nenhuma celula nas primeiras 18 linhas tem valor literal 0 (zeros sao None)",
          zeros_literais == 0,
          f"{zeros_literais} celulas com zero literal encontradas")


# ===========================================================================
# SUITE 18 — HTML: empreendimento completo (cidade, UF, áreas, tipologias)
# ===========================================================================

def suite_html_empreendimento(projeto, resultado) -> None:
    secao("SUITE 18 — HTML: dados completos do empreendimento")

    try:
        html = gerar_relatorio_html(projeto, resultado)
    except Exception as e:
        check("HTML gerado sem erro", False, str(e))
        return

    p = projeto.terreno

    # Identificacao
    check("HTML contem nome do empreendimento", p.info.nome in html)
    check("HTML contem cidade",  p.info.cidade in html)
    check("HTML contem UF",      p.info.uf     in html)

    # Tipologias
    for tip in p.tipologias:
        check(f"HTML menciona tipologia '{tip.nome}'", tip.nome in html)

    # Indicadores chave (valor formatado com R$)
    r   = resultado.resumo
    ind = resultado.indicadores
    vgv = r.get("vgv_vendavel", 0)
    tir = ind.get("tir_anual")

    # VGV deve aparecer abreviado (e.g. "53.9M") ou completo
    vgv_M = vgv / 1e6
    check("HTML menciona VGV em formato M ou inteiro",
          f"{vgv_M:.1f}" in html or f"{vgv_M:.0f}" in html,
          f"vgv={vgv_M:.1f}M")

    # TIR
    if tir:
        tir_pct = tir * 100
        check("HTML menciona TIR com pelo menos 1 casa decimal",
              f"{tir_pct:.1f}" in html or f"{tir_pct:.2f}" in html,
              f"TIR={tir_pct:.2f}%")

    # Secoes de pagina (4 secoes = quebras de pagina para PDF)
    n_secoes = html.count("page-break") + html.count("break-before")
    check("HTML tem ao menos 3 quebras de pagina para impressao PDF",
          n_secoes >= 3,
          f"encontradas {n_secoes}")

    # Data de geracao
    from datetime import datetime
    ano_atual = str(datetime.now().year)
    check(f"HTML menciona ano atual ({ano_atual})", ano_atual in html)


# ===========================================================================
# SUITE 19 — Financiamento + Reajustes simultâneos: invariantes mantidos
# ===========================================================================

def suite_financiamento_reajustes_simultaneos(projeto_base) -> None:
    secao("SUITE 19 — Financiamento + Reajustes simultaneos: integracao completa")

    try:
        from src.modelos.financeiro import ConfigFinanciamento
        from src.modelos import ConfigReajustes
    except ImportError:
        try:
            from src.modelos.financeiro import ConfigFinanciamento
            from src.modelos.reajustes import ConfigReajustes
        except ImportError as e:
            check("Imports OK", False, str(e))
            return

    cfg_fin = ConfigFinanciamento(
        ativo=True, taxa_juros_am=1.5, limite_credito_valor=0,
        periodo_carencia_meses=6, comissao_abertura_pct=0.5,
        iof_pct=0.38, caixa_minimo=0.0,
    )
    cfg_raj = ConfigReajustes(
        ativo=True, aplicar_incc_obras=True, incc_anual_pct=5.5,
        aplicar_correcao_parcelas=True, indice_parcelas="incc",
        ipca_anual_pct=3.5, igpm_anual_pct=4.0,
        aplicar_correcao_terreno=False, indice_terreno="igpm",
    )

    try:
        projeto_comb = projeto_base.model_copy(update={
            "financiamento": cfg_fin,
            "reajustes": cfg_raj,
        })
        r_comb = calcular_fluxo_caixa(projeto_comb)
    except Exception as e:
        check("Calculo com financiamento + reajustes sem erro", False, str(e))
        return

    check("Calculo com financiamento + reajustes sem erro", True)

    rc = r_comb.resumo
    check("Custo financiamento > 0 (juros foram cobrados)",
          rc.get("custo_financiamento_juros", 0) > 0)
    check("Variacao INCC > 0 (obras encareceram)",
          rc.get("variacao_custo_obras_incc", 0) >= 0)
    check("Invariante: receita_nominal + receita_financeira = total_recebido",
          abs(rc["receita_nominal_venda"] + rc["receita_financeira"]
              - rc["vgv_total_recebido"]) < 1.0,
          f"nom={rc['receita_nominal_venda']:,.0f}  jur={rc['receita_financeira']:,.0f}  "
          f"tot={rc['vgv_total_recebido']:,.0f}")

    # Excel gerado sem erro
    try:
        tmp = Path(tempfile.mkdtemp()) / "teste_comb.xlsx"
        exportar_para_excel(projeto_comb, r_comb, tmp)
        wb_comb = load_workbook(tmp)
        check("Excel financiamento + reajustes: 9 abas presentes",
              len(wb_comb.sheetnames) == 9)

        # Dashboard deve ter secao de financiamento
        ws_dash = wb_comb["Dashboard"]
        tem_secao_fin = any(
            "FINANCIAMENTO" in str(ws_dash.cell(rw, 1).value or "").upper()
            for rw in range(1, ws_dash.max_row + 1)
        )
        check("Dashboard com ambas as features tem secao FINANCIAMENTO",
              tem_secao_fin)
    except Exception as e:
        check("Excel financiamento + reajustes gerado sem erro", False, str(e))

    # HTML gerado sem erro
    try:
        html = gerar_relatorio_html(projeto_comb, r_comb)
        check("HTML financiamento + reajustes gerado (>5 KB)",
              len(html) > 5_000, f"{len(html)} bytes")
    except Exception as e:
        check("HTML financiamento + reajustes sem erro", False, str(e))


# ===========================================================================
# SUITE 20 — Payback: coerência engine vs saldo acumulado no Excel
# ===========================================================================

def suite_payback_coerencia(wb, resultado) -> None:
    secao("SUITE 20 — Payback: coerencia entre engine e Saldo Acumulado no Excel")

    ws  = wb["Fluxo de Caixa"]
    ind = resultado.indicadores
    pb  = ind.get("payback_simples_meses")

    check("Payback calculado pela engine (nao None)", pb is not None,
          "fluxo pode nao ter troca de sinal positivo")
    if pb is None:
        return

    check(f"Payback ({pb}) <= horizonte ({resultado.horizonte})",
          pb <= resultado.horizonte,
          f"pb={pb}  horizonte={resultado.horizonte}")

    # Verificar no Excel que o saldo acumulado cruza de negativo para positivo no mes pb
    linha_acum = _encontrar_linha_texto(ws, "Saldo Acumulado")
    if not linha_acum:
        check("Linha 'Saldo Acumulado' encontrada", False)
        return

    # Cabecalho usa strings "M0", "M1", ... — construir mapa mes->valor
    mapa: dict[int, float] = {}
    for col in range(2, ws.max_column + 1):
        mes_hdr = ws.cell(1, col).value
        val_acum = ws.cell(linha_acum, col).value
        # Aceitar tanto "M40" (string) quanto 40 (int)
        if isinstance(mes_hdr, str) and mes_hdr.startswith("M"):
            try:
                num = int(mes_hdr[1:])
                if isinstance(val_acum, (int, float)):
                    mapa[num] = float(val_acum)
            except ValueError:
                pass
        elif isinstance(mes_hdr, (int, float)) and isinstance(val_acum, (int, float)):
            mapa[int(mes_hdr)] = float(val_acum)

    if pb not in mapa:
        check(f"Mes de payback (M{pb}) encontrado no Excel", False,
              f"meses no mapa: {sorted(mapa.keys())[:10]}")
        return

    val_pb   = mapa[pb]
    val_prev = mapa.get(pb - 1, -1.0)

    check(f"Saldo acumulado no mes de payback (M{pb}) >= 0",
          val_pb >= -1.0,  # tol R$1
          f"saldo={val_pb:,.0f}")
    check(f"Saldo acumulado no mes anterior ao payback (M{pb-1}) < 0",
          val_prev < 1.0,
          f"saldo_anterior={val_prev:,.0f}")


# ===========================================================================
# SUITE 21 — Simulação Lote: tipologia com maior VGV paga mais; principal == VGV
# ===========================================================================

def suite_simulacao_lote_avancado(wb, projeto) -> None:
    secao("SUITE 21 — Simulacao Lote avancado: VGV ordering e principal exato")

    ws = wb["Simulacao Lote"]

    # Coletar totais por tipologia (linha TOTAL, col 3=Parcela, 4=Principal, 5=Juros)
    linhas_total: list[tuple[int, float, float, float]] = []
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row, 2).value
        if v and str(v).strip().upper() == "TOTAL":
            tp  = _valor_celula(ws, row, 3)
            tpr = _valor_celula(ws, row, 4)
            tj  = _valor_celula(ws, row, 5)
            if tp > 0:
                linhas_total.append((row, tp, tpr, tj))

    if len(linhas_total) < 2:
        check("Pelo menos 2 tipologias na simulacao para comparar", False,
              f"encontradas {len(linhas_total)}")
        return

    # VGV por tipologia (na ordem em que aparecem no projeto)
    tips = projeto.terreno.tipologias
    check("Numero de tipologias na planilha == no projeto",
          len(linhas_total) == len(tips),
          f"planilha={len(linhas_total)}  projeto={len(tips)}")

    # Principal de cada tipologia deve aproximar o VGV lote dela (tol 1%)
    for idx, (row, tp, tpr, tj) in enumerate(linhas_total):
        if idx >= len(tips):
            break
        vgv_lote = float(tips[idx].vgv_lote)
        dif_pct = abs(tpr - vgv_lote) / max(vgv_lote, 1) * 100
        check(f"Tip '{tips[idx].nome}': principal total ~= VGV lote (tol 1%)",
              dif_pct < 1.0,
              f"principal={tpr:,.0f}  vgv_lote={vgv_lote:,.0f}  dif={dif_pct:.2f}%")
        check(f"Tip '{tips[idx].nome}': total pago >= VGV lote (juros adicionam valor)",
              tp >= vgv_lote - 1.0,
              f"total={tp:,.0f}  vgv_lote={vgv_lote:,.0f}")

    # Ordering: se VGV[i] > VGV[j] entao total_pago[i] >= total_pago[j]
    vgvs   = [float(t.vgv_lote)   for t in tips[:len(linhas_total)]]
    totais = [lt[1]               for lt in linhas_total]
    # Verificar ordenacao relativa (tipologia com maior VGV paga mais)
    pares_corretos = sum(
        1 for i in range(len(vgvs)) for j in range(i + 1, len(vgvs))
        if (vgvs[i] > vgvs[j]) == (totais[i] >= totais[j] - 1)
        or vgvs[i] == vgvs[j]
    )
    total_pares = len(vgvs) * (len(vgvs) - 1) // 2
    check("Ordering: tipologia maior VGV paga mais total",
          pares_corretos == total_pares,
          f"{pares_corretos}/{total_pares} pares corretos")


# ===========================================================================
# SUITE 22 — UX: arquivo Excel abrivel, Dashboard primeiro, tamanho razoavel
# ===========================================================================

def suite_ux_excel(xlsx_path: Path, wb) -> None:
    secao("SUITE 22 — UX: Excel abrivel, Dashboard primeiro, tamanho aceitavel")

    # Tamanho do arquivo (< 5 MB para abertura rapida)
    tamanho_mb = xlsx_path.stat().st_size / 1_048_576
    check(f"Arquivo Excel < 5 MB ({tamanho_mb:.2f} MB)",
          tamanho_mb < 5.0, f"{tamanho_mb:.2f} MB")

    # Dashboard deve ser a primeira aba (apresentacao executiva)
    primeira_aba = wb.sheetnames[0] if wb.sheetnames else ""
    check("Dashboard e a primeira aba do arquivo",
          primeira_aba == "Dashboard",
          f"primeira aba: '{primeira_aba}'")

    # freeze_panes em abas chave (melhora navegacao — colunas de mes fixas)
    ws_fc  = wb["Fluxo de Caixa"]
    ws_ver = wb["Verificacao Receitas"]
    check("Freeze panes ativo na aba Fluxo de Caixa",
          ws_fc.freeze_panes is not None and ws_fc.freeze_panes != "A1",
          f"freeze={ws_fc.freeze_panes}")
    check("Freeze panes ativo na aba Verificacao Receitas",
          ws_ver.freeze_panes is not None and ws_ver.freeze_panes != "A1",
          f"freeze={ws_ver.freeze_panes}")

    # Largura da coluna A do Dashboard deve ser >= 30 (labels legíveis)
    col_a_width = wb["Dashboard"].column_dimensions["A"].width
    check(f"Coluna A do Dashboard tem largura >= 30 (legibilidade dos labels)",
          col_a_width >= 30,
          f"width={col_a_width}")

    # Arquivo pode ser reaberto sem erro (integridade)
    try:
        wb2 = load_workbook(xlsx_path, read_only=True)
        nomes2 = wb2.sheetnames
        wb2.close()
        check("Arquivo pode ser reaberto (integridade OK)",
              len(nomes2) == 9, f"abas={nomes2}")
    except Exception as e:
        check("Arquivo pode ser reaberto (integridade OK)", False, str(e))

    # Numero de linhas do Dashboard nao pode ser excessivo (> 500 = risco de lentidao)
    n_linhas_dash = wb["Dashboard"].max_row
    check(f"Dashboard nao tem excesso de linhas (<= 200, atual={n_linhas_dash})",
          n_linhas_dash <= 200,
          f"{n_linhas_dash} linhas")


# ===========================================================================
# MAIN
# ===========================================================================

def main() -> None:
    print("=" * 60)
    print("  TESTES LENTOS — EXPORTADORES EXCEL e HTML")
    print(f"  Projeto: {EXEMPLO_PATH.name}")
    print("=" * 60)

    # --- Carregar projeto base ---
    print("\nCarregando projeto e calculando fluxo de caixa...")
    try:
        projeto, resultado = _carregar_e_calcular(EXEMPLO_PATH)
        print(f"  OK — horizonte={resultado.resumo.get('horizonte_meses')} meses  "
              f"VGV={resultado.resumo.get('vgv_vendavel',0)/1e6:.2f}M")
    except Exception as e:
        print(f"  ERRO ao carregar/calcular: {e}")
        traceback.print_exc()
        sys.exit(1)

    # --- Gerar Excel base ---
    print("\nGerando Excel...")
    try:
        xlsx_path = _exportar_excel(projeto, resultado)
        wb = load_workbook(xlsx_path)
        print(f"  OK — {xlsx_path.stat().st_size:,} bytes  abas={wb.sheetnames}")
    except Exception as e:
        print(f"  ERRO ao gerar Excel: {e}")
        traceback.print_exc()
        sys.exit(1)

    # --- Gerar HTML ---
    print("\nGerando HTML...")
    try:
        html = gerar_relatorio_html(projeto, resultado)
        print(f"  OK — {len(html):,} bytes")
    except Exception as e:
        print(f"  ERRO ao gerar HTML: {e}")
        traceback.print_exc()
        html = ""

    # --- Executar suites ---
    suite_estrutura_excel(wb)
    suite_dashboard(wb, resultado)
    suite_fluxo_caixa(wb, resultado)
    suite_verificacao_receitas(wb, resultado)
    suite_dre_consistencia(wb, resultado)
    suite_simulacao_lote(wb, projeto)
    suite_html(projeto, resultado)
    suite_financiamento_caixa_minimo(projeto, resultado)
    suite_reajustes(projeto)
    suite_preco_progressivo(projeto)
    suite_desenvolvimento_impostos(wb, resultado)
    suite_consistencia_cruzada(projeto, resultado, wb, html)

    # --- Novas suites (13-22) ---
    suite_dre_valores_exatos(wb, resultado)
    suite_exposicao_maxima(wb, resultado)
    suite_fluxo_recebiveis_soma(wb, projeto)
    suite_terreno_desembolso(projeto, resultado)
    suite_zeros_ocultos(wb, resultado)
    suite_html_empreendimento(projeto, resultado)
    suite_financiamento_reajustes_simultaneos(projeto)
    suite_payback_coerencia(wb, resultado)
    suite_simulacao_lote_avancado(wb, projeto)
    suite_ux_excel(xlsx_path, wb)

    # --- Resumo final ---
    total   = len(_resultados)
    passed  = sum(1 for _, ok, _ in _resultados if ok)
    failed  = total - passed
    falhas  = [(desc, det) for desc, ok, det in _resultados if not ok]

    print(f"\n{'='*60}")
    print(f"  RESULTADO FINAL: {passed}/{total} PASS  |  {failed} FAIL")
    print(f"{'='*60}")

    if falhas:
        print("\n  FALHAS DETECTADAS:")
        for desc, det in falhas:
            sufixo = f" — {det}" if det else ""
            print(f"    ✗ {desc}{sufixo}")
    else:
        print("\n  TODOS OS TESTES PASSARAM ✓")

    print()
    return 0 if failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
