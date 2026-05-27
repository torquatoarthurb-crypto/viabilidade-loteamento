"""
Exportacao de resultado para Excel — ferramenta de auditoria completa.

Abas:
  1. Dashboard        — KPIs, DRE, composicao custos (apresentacao executiva)
  2. Terreno          — inputs do empreendimento + receita mensal
  3. Receitas         — fluxos recebiveis + curva de vendas + receita mensal
  4. Obras            — etapas + desembolso mensal por etapa
  5. Desenvolvimento  — despesas + desembolso mensal por item
  6. Impostos         — regime tributario + comissao + fluxo mensal
  7. Fluxo de Caixa   — tabela mensal completa (todas as categorias)
  8. Verificacao Rec. — detalhe de recebimentos e invariantes matematicos
  9. Simulacao Lote   — cronograma de pagamento de 1 lote por tipologia

Convencao de zeros: valores zero exibidos como celulas vazias via number_format
com terceiro segmento vazio: 'R$ #,##0;[Red]-R$ #,##0;""'
"""

from __future__ import annotations

from datetime import date as _date
from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..engine import ResultadoCalculo, simular_lote_unitario
from ..engine.despesas import _distribuir_etapa, _distribuir_despesa_temporal
from ..engine.recebimentos import calcular_vgv_vendavel
from ..engine.utilidades import meses_entre, parcela_price
from ..modelos import Projeto


# =====================================================================
# PALETA ALTIPLANO
# =====================================================================

_C_DARK    = "1A1916"   # carvão Altiplano
_C_OCHRE   = "B07D2E"   # ocre Altiplano — seções principais
_C_OCHRE_L = "F5EED8"   # ocre claro — subseções
_C_STONE   = "F5F3EE"   # pedra clara
_C_STONE_D = "ECEAE4"   # pedra média
_C_FLOW_H  = "2C2A27"   # carvão escuro — cabeçalho do fluxo
_C_GREEN   = "D4EAD9"   # verde suave — resultado positivo
_C_RED     = "F5E8E8"   # vermelho suave — resultado negativo
_C_WHITE   = "FFFFFF"

_F_W11  = Font(name="Calibri", size=11, bold=True,  color="FFFFFF")
_F_W10  = Font(name="Calibri", size=10, bold=True,  color="FFFFFF")
_F_B10  = Font(name="Calibri", size=10, bold=True,  color="1A1916")
_F_N10  = Font(name="Calibri", size=10,              color="1A1916")
_F_S9   = Font(name="Calibri", size=9,               color="5A5650")

# Formatos de numero — zeros ficam em branco (terceiro segmento: "")
_FMT_RS   = 'R$ #,##0;[Red]-R$ #,##0;""'
_FMT_RS2  = 'R$ #,##0.00;[Red]-R$ #,##0.00;""'
_FMT_PCT  = '0.0%;[Red]-0.0%;""'
_FMT_PCT2 = '0.00%;[Red]-0.00%;""'
_FMT_INT  = '#,##0;-#,##0;""'
_FMT_NUM  = '#,##0.00;-#,##0.00;""'
_FMT_MES  = '0'

# Layout das colunas nas abas de auditoria
_C_IN_LABEL = 1   # A — label do input
_C_IN_VAL   = 2   # B — valor do input
_C_IN_UNIT  = 3   # C — unidade / obs
_C_GAP      = 4   # D — separador visual
_C_MES      = 5   # E — numero do mes (fluxo)
_C_FLOW1    = 6   # F — primeira coluna do fluxo


# =====================================================================
# HELPERS BASICOS
# =====================================================================

def _fill(cor: str) -> PatternFill:
    return PatternFill("solid", fgColor=cor)


def _hdr(ws, row: int, col: int, text: str, cor: str,
         fonte=None, ncols: int = 1, height: int = 20) -> None:
    cell = ws.cell(row, col, text)
    cell.font = fonte or _F_W10
    cell.fill = _fill(cor)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = height
    if ncols > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + ncols - 1,
        )


def _kv(ws, row: int, label: str, val=None, fmt: str | None = None,
        bold: bool = False, cl: int = _C_IN_LABEL, cv: int = _C_IN_VAL,
        unit: str = "", cor_val: str | None = None) -> int:
    """Escreve par label/valor. Retorna proxima linha."""
    if label == "" and val is None:
        return row + 1
    c_label = ws.cell(row, cl, label)
    c_label.font = _F_B10 if bold else _F_N10
    if val is not None and val != "":
        c_val = ws.cell(row, cv, val)
        c_val.font = _F_B10 if bold else _F_N10
        c_val.alignment = Alignment(horizontal="right")
        if fmt:
            c_val.number_format = fmt
        if cor_val:
            c_val.fill = _fill(cor_val)
    if unit:
        c_unit = ws.cell(row, cl + 2, unit)
        c_unit.font = _F_S9
    return row + 1


def _sec(ws, row: int, titulo: str, ncols: int = 3, altura: int = 18) -> int:
    """Cabecalho de secao ocre."""
    _hdr(ws, row, 1, titulo, _C_OCHRE, fonte=_F_W10, ncols=ncols, height=altura)
    return row + 1


def _sub(ws, row: int, titulo: str, ncols: int = 3) -> int:
    """Sub-cabecalho ocre claro."""
    _hdr(ws, row, 1, titulo, _C_OCHRE_L, fonte=_F_B10, ncols=ncols, height=16)
    return row + 1


def _dim_auditoria(ws) -> None:
    """Define larguras das colunas no layout de auditoria."""
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 2
    ws.column_dimensions["E"].width = 7


def _titulo_projeto(ws, projeto: Projeto, row: int = 1) -> int:
    """Linha de titulo do projeto no topo da aba."""
    _hdr(ws, row, 1, f"{projeto.terreno.info.nome}  —  {projeto.terreno.info.cidade}/{projeto.terreno.info.uf}",
         _C_DARK, fonte=_F_W11, ncols=40, height=24)
    return row + 2


# =====================================================================
# HELPERS DE FLUXO (lado direito das abas de auditoria)
# =====================================================================

def _cabecalho_fluxo(ws, row: int, colunas: list[str], largura: int = 15) -> int:
    """Escreve o cabecalho da secao de fluxo (colunas = nomes das series)."""
    _hdr(ws, row, _C_GAP, "FLUXO MENSAL", _C_FLOW_H,
         fonte=_F_W10, ncols=2 + len(colunas))
    # Mes
    c_mes = ws.cell(row + 1, _C_MES, "Mes")
    c_mes.font = _F_B10
    c_mes.fill = _fill(_C_STONE_D)
    c_mes.alignment = Alignment(horizontal="center")
    # Colunas de serie
    for i, nome in enumerate(colunas):
        col = _C_FLOW1 + i
        c = ws.cell(row + 1, col, nome)
        c.font = _F_B10
        c.fill = _fill(_C_STONE_D)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        if col > _C_FLOW1:
            ws.column_dimensions[get_column_letter(col)].width = largura
    ws.column_dimensions[get_column_letter(_C_MES)].width = 7
    ws.column_dimensions[get_column_letter(_C_FLOW1)].width = largura
    return row + 2


def _linhas_fluxo(ws, row_ini: int, meses: list[int],
                  series: list[list[float]], fmt: str = _FMT_RS,
                  total_col: bool = False) -> None:
    """Preenche as linhas mensais do fluxo. series[i] = lista de valores por mes."""
    n_series = len(series)
    for idx, mes in enumerate(meses):
        row = row_ini + idx
        # Mes
        mes_int = int(mes)
        cm = ws.cell(row, _C_MES, f"M{mes_int}")
        cm.font = _F_S9
        cm.alignment = Alignment(horizontal="center")
        cm.fill = _fill(_C_STONE if mes_int % 2 == 0 else _C_WHITE)
        # Valores
        for si in range(n_series):
            col = _C_FLOW1 + si
            val = series[si][idx] if idx < len(series[si]) else 0.0
            cv = ws.cell(row, col, val if val != 0.0 else None)
            cv.number_format = fmt
            cv.font = _F_N10
            cv.fill = _fill(_C_STONE if mes % 2 == 0 else _C_WHITE)
            cv.alignment = Alignment(horizontal="right")


# =====================================================================
# 1. DASHBOARD
# =====================================================================

def _dashboard(wb: Workbook, projeto: Projeto, resultado: ResultadoCalculo) -> None:
    ws = wb.create_sheet("Dashboard")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 16
    ws.freeze_panes = "A3"

    r  = resultado.resumo
    ind = resultado.indicadores
    p  = projeto.terreno
    tma = projeto.parametros.tma_anual

    def _fmt_rs(v) -> str:
        if v is None:
            return "—"
        n = int(round(abs(v)))
        s = f"{n:,}".replace(",", ".")
        return f"R$ {s}" if v >= 0 else f"-R$ {s}"

    def _fmt_pct(v) -> str:
        return f"{v * 100:.1f}%" if v is not None else "—"

    linha = 1

    # Titulo
    _hdr(ws, linha, 1, "ANALISE DE VIABILIDADE ECONOMICA", _C_DARK,
         fonte=Font(name="Calibri", size=16, bold=True, color="FFFFFF"),
         ncols=3, height=30)
    linha += 1
    _hdr(ws, linha, 1, f"{p.info.nome}  —  {p.info.cidade}/{p.info.uf}  |  {_date.today().strftime('%d/%m/%Y')}",
         _C_OCHRE, fonte=_F_W11, ncols=3, height=20)
    linha += 2

    # KPIs
    linha = _sec(ws, linha, "INDICADORES PRINCIPAIS", ncols=3, altura=20)

    tir = ind.get("tir_anual")
    pb  = ind.get("payback_simples_meses")
    mult = ind.get("lucro_sobre_exposicao")
    margem_vv = r.get("margem_sobre_vgv_vendavel")
    margem_vb = r.get("margem_sobre_vgv_bruto")

    kpis = [
        ("Horizonte do projeto",         f"{r.get('horizonte_meses', 0)} meses",   False, None),
        ("Data de inicio (M0)",          p.datas.inicio_projeto.strftime("%m/%Y"), False, None),
        ("Total de lotes",               f"{p.total_lotes} lotes",                 False, None),
        ("", "", False, None),
        ("VGV Bruto",                    _fmt_rs(r["vgv_bruto"]),    False, None),
        ("VGV Vendavel",                 _fmt_rs(r["vgv_vendavel"]), False, None),
        ("Receita Total Recebida",       _fmt_rs(r["vgv_total_recebido"]), False, None),
        ("", "", False, None),
        ("Resultado Bruto (Lucro)",      _fmt_rs(r["lucro_liquido"]), True,
         _C_GREEN if r["lucro_liquido"] >= 0 else _C_RED),
        ("Margem s/ VGV Vendavel",       _fmt_pct(margem_vv), True,
         _C_GREEN if (margem_vv or 0) >= 0 else _C_RED),
        ("Margem s/ VGV Bruto",         _fmt_pct(margem_vb), False, None),
        ("", "", False, None),
        ("TIR (ao ano)",                 f"{tir*100:.2f}%" if tir else "—", True, None),
        *([("  TIR sem financiamento",   f"{r['tir_anual_sem_fin']*100:.2f}%", False, None)]
          if projeto.financiamento.ativo and r.get("tir_anual_sem_fin") else []),
        (f"VPL (TMA {tma:.0f}% a.a.)",  _fmt_rs(ind["vpl"]), True,
         _C_GREEN if ind["vpl"] >= 0 else _C_RED),
        ("Payback Simples",              f"Mes {pb}" if pb else "—", False, None),
        ("Exposicao Maxima de Caixa",   _fmt_rs(ind["exposicao_maxima"]), False, _C_RED),
        ("Multiplicador Lucro/Exposicao", f"{mult:.2f}x".replace(".", ",") if mult else "—", False, None),
    ]

    for label, val, negrito, cor_fundo in kpis:
        if not label:
            linha += 1
            continue
        cl = ws.cell(linha, 1, label)
        cl.font = _F_B10 if negrito else _F_N10
        if cor_fundo:
            cl.fill = _fill(cor_fundo)
        cv = ws.cell(linha, 2, val)
        cv.font = _F_B10 if negrito else _F_N10
        cv.alignment = Alignment(horizontal="right")
        if cor_fundo:
            cv.fill = _fill(cor_fundo)
        linha += 1

    linha += 1

    # DRE
    linha = _sec(ws, linha, "DEMONSTRATIVO DE RESULTADO (DRE)", ncols=3, altura=20)

    vgv_efetivo = r.get("vgv_efetivo_vendavel", r["vgv_vendavel"])
    ajuste_preco = vgv_efetivo - r["vgv_vendavel"]

    dre = [
        ("VGV Bruto (preco base)",                  r["vgv_bruto"],          False, _C_STONE),
        ("  (-) Permuta Fisica / Deducoes",         -(r["vgv_bruto"] - r["vgv_vendavel"]), False, _C_WHITE),
        ("VGV Vendavel (preco base)",                r["vgv_vendavel"],        True,  _C_STONE),
        ("  (+/-) Ajuste Preco Progressivo",        ajuste_preco if abs(ajuste_preco) > 0.01 else None, False, _C_WHITE),
        ("VGV Efetivo Vendavel",                    vgv_efetivo if abs(ajuste_preco) > 0.01 else None, True, _C_STONE),
        ("  (+) Receita Financeira (Juros)",         r["receita_financeira"],  False, _C_WHITE),
        ("Receita Total Recebida",                   r["vgv_total_recebido"],  True,  _C_STONE),
        ("",                                         None,                     False, _C_WHITE),
        ("  (-) Aquisicao do Terreno",              -r["custo_terreno_aquisicao"], False, _C_WHITE),
        ("  (-) Cartorio",                           -r["custo_terreno_cartorio"], False, _C_WHITE),
        ("  (-) Obras (com BDI e Contingencia)",    -r["custo_obras"],         False, _C_WHITE),
        ("  (-) Projetos",                          -r["custo_projetos"],      False, _C_WHITE),
        ("  (-) Licenciamento",                     -r["custo_licenciamento"], False, _C_WHITE),
        ("  (-) Marketing",                         -r["custo_marketing"],     False, _C_WHITE),
        ("  (-) Outros Desenvolvimento",            -r["custo_outros"],        False, _C_WHITE),
        ("  (-) Administracao",                     -r["custo_administracao"], False, _C_WHITE),
        ("  (-) Comissao de Vendas",                -r["custo_comissao"],      False, _C_WHITE),
        ("  (-) Impostos / Tributacao",             -r["custo_impostos"],      False, _C_WHITE),
        ("  (-) Permuta Financeira",                -r["custo_permuta_financeira"], False, _C_WHITE),
        ("Total de Saidas",                         -r["total_saidas"],         True,  _C_STONE),
        ("",                                        None,                       False, _C_WHITE),
        ("RESULTADO BRUTO (LUCRO LIQUIDO)",         r["lucro_liquido"],         True,
         _C_GREEN if r["lucro_liquido"] >= 0 else _C_RED),
    ]

    vgv_b = r["vgv_bruto"] if r["vgv_bruto"] > 0 else 1.0
    ws.cell(linha, 1, "Item").font = _F_B10
    ws.cell(linha, 1).fill = _fill(_C_STONE_D)
    ws.cell(linha, 2, "Valor (R$)").font = _F_B10
    ws.cell(linha, 2).fill = _fill(_C_STONE_D)
    ws.cell(linha, 2).alignment = Alignment(horizontal="right")
    ws.cell(linha, 3, "% VGV Bruto").font = _F_B10
    ws.cell(linha, 3).fill = _fill(_C_STONE_D)
    ws.cell(linha, 3).alignment = Alignment(horizontal="right")
    linha += 1

    for label, val, negrito, cor in dre:
        if not label:
            linha += 1
            continue
        cl = ws.cell(linha, 1, label)
        cl.font = _F_B10 if negrito else _F_N10
        cl.fill = _fill(cor)
        if val is not None:
            cv = ws.cell(linha, 2, val)
            cv.font = _F_B10 if negrito else _F_N10
            cv.number_format = _FMT_RS
            cv.alignment = Alignment(horizontal="right")
            cv.fill = _fill(cor)
            cp = ws.cell(linha, 3, abs(val) / vgv_b if vgv_b else 0)
            cp.font = _F_N10
            cp.number_format = _FMT_PCT
            cp.alignment = Alignment(horizontal="right")
            cp.fill = _fill(cor)
        linha += 1

    # Secao de financiamento bancario (so quando ativo)
    if projeto.financiamento.ativo and r.get("custo_financiamento_total"):
        linha += 1
        linha = _sec(ws, linha, "FINANCIAMENTO BANCARIO (CCB/CCE)", ncols=3, altura=18)
        fin = projeto.financiamento
        itens_fin = [
            ("Taxa de juros",               f"{fin.taxa_juros_am:.2f}% a.m.  /  {(1+fin.taxa_juros_am/100)**12-1:.2f}% a.a.", False),
            ("Juros pagos ao banco",        _fmt_rs(r.get("custo_financiamento_juros",  0)), False),
            ("Comissao de abertura",         _fmt_rs(r.get("custo_financiamento_comissao", 0)), False),
            ("Custo total do financiamento", _fmt_rs(r.get("custo_financiamento_total",  0)), True),
            ("Saldo devedor maximo",         _fmt_rs(r.get("saldo_devedor_maximo",        0)), False),
            ("Saldo devedor remanescente",   _fmt_rs(r.get("saldo_devedor_final",         0)),
             r.get("saldo_devedor_final", 0) > 1),
        ]
        limite_cred = fin.limite_credito_valor
        saldo_max_v = r.get("saldo_devedor_maximo", 0) or 0
        for label, val, negrito in itens_fin:
            cl = ws.cell(linha, 1, label)
            cl.font = _F_B10 if negrito else _F_N10
            cv = ws.cell(linha, 2, val)
            cv.font = _F_B10 if negrito else _F_N10
            cv.alignment = Alignment(horizontal="right")
            if label.startswith("Saldo devedor rem") and r.get("saldo_devedor_final", 0) > 1:
                cv.fill = _fill(_C_RED)
                cl.fill = _fill(_C_RED)
            elif label.startswith("Saldo devedor max") and limite_cred > 0 and saldo_max_v > limite_cred:
                cv.fill = _fill(_C_RED)
                cl.fill = _fill(_C_RED)
            linha += 1

    # Secao de investidor (so quando ativo)
    if r.get("investidor_ativo"):
        linha += 1
        modo_inv = r.get("investidor_modo", "pct_negocio")
        fin = projeto.financiamento
        if modo_inv == "pct_negocio":
            linha = _sec(ws, linha, "INVESTIDOR — % DO NEGÓCIO", ncols=3, altura=18)
            pct_inv = r.get("investidor_pct_negocio", 0)
            lucro_inv = r.get("lucro_investidor", 0)
            lucro_lot = r.get("lucro_loteadora", 0)
            for label, val, negrito, cor_f in [
                ("Participacao do investidor",     f"{pct_inv:.1f}%",  False, None),
                ("Lucro destinado ao investidor",  _fmt_rs(lucro_inv), False, None),
                ("Lucro remanescente (loteadora)", _fmt_rs(lucro_lot), True,
                 _C_GREEN if lucro_lot >= 0 else _C_RED),
            ]:
                cl = ws.cell(linha, 1, label)
                cl.font = _F_B10 if negrito else _F_N10
                cv = ws.cell(linha, 2, val)
                cv.font = _F_B10 if negrito else _F_N10
                cv.alignment = Alignment(horizontal="right")
                if cor_f:
                    cl.fill = _fill(cor_f)
                    cv.fill = _fill(cor_f)
                linha += 1
        else:
            linha = _sec(ws, linha, "INVESTIDOR — EMPRÉSTIMO", ncols=3, altura=18)
            custo_j_inv = r.get("custo_juros_investidor", 0) or 0
            saldo_max_i = r.get("saldo_devedor_investidor_maximo", 0) or 0
            saldo_fin_i = r.get("saldo_devedor_investidor_final", 0) or 0
            for label, val, negrito, cor_f in [
                ("Taxa de juros investidor",      f"{fin.taxa_juros_investidor_am:.2f}% a.m.", False, None),
                ("Juros pagos ao investidor",     _fmt_rs(custo_j_inv), False, None),
                ("Saldo devedor maximo",           _fmt_rs(saldo_max_i), False, None),
                ("Saldo devedor remanescente",     _fmt_rs(saldo_fin_i), saldo_fin_i > 1,
                 _C_RED if saldo_fin_i > 1 else None),
            ]:
                cl = ws.cell(linha, 1, label)
                cl.font = _F_B10 if negrito else _F_N10
                cv = ws.cell(linha, 2, val)
                cv.font = _F_B10 if negrito else _F_N10
                cv.alignment = Alignment(horizontal="right")
                if cor_f:
                    cl.fill = _fill(cor_f)
                    cv.fill = _fill(cor_f)
                linha += 1

    # Secao de reajustes monetarios (so quando ativo)
    if projeto.reajustes.ativo and (r.get("variacao_custo_obras_incc") or r.get("receita_correcao_total")):
        linha += 1
        linha = _sec(ws, linha, "REAJUSTES MONETARIOS", ncols=3, altura=18)
        reaj      = projeto.reajustes
        var_incc  = r.get("variacao_custo_obras_incc", 0) or 0
        corr_parc = r.get("receita_correcao_total",    0) or 0
        efeito    = corr_parc - var_incc
        itens_reaj = [
            ("INCC obras projetado",         f"{reaj.incc_anual_pct:.2f}% a.a." if reaj.aplicar_incc_obras else "—", False),
            ("Variacao custo obras (INCC)",  _fmt_rs(var_incc)  if var_incc  else "—", False),
            ("Correcao recebida (parcelas)", _fmt_rs(corr_parc) if corr_parc else "—", False),
            ("Efeito liquido no resultado",  _fmt_rs(efeito),   True),
        ]
        for label, val, negrito in itens_reaj:
            cl = ws.cell(linha, 1, label)
            cl.font = _F_B10 if negrito else _F_N10
            cv = ws.cell(linha, 2, val)
            cv.font = _F_B10 if negrito else _F_N10
            cv.alignment = Alignment(horizontal="right")
            if negrito:
                cv.fill = _fill(_C_GREEN if efeito >= 0 else _C_RED)
            linha += 1


# =====================================================================
# 2. TERRENO
# =====================================================================

def _aba_terreno(wb: Workbook, projeto: Projeto, resultado: ResultadoCalculo) -> None:
    ws = wb.create_sheet("Terreno")
    _dim_auditoria(ws)
    ws.freeze_panes = "E5"

    p  = projeto.terreno
    df = resultado.fluxo_caixa
    linha = _titulo_projeto(ws, projeto)

    # ---- INPUTS ----
    linha = _sec(ws, linha, "IDENTIFICACAO", ncols=3)
    linha = _kv(ws, linha, "Nome do empreendimento", p.info.nome)
    linha = _kv(ws, linha, "Cidade / UF", f"{p.info.cidade} / {p.info.uf}")
    linha += 1

    linha = _sec(ws, linha, "QUADRO DE AREAS", ncols=3)
    linha = _kv(ws, linha, "Area da gleba",            p.areas.area_gleba_m2,          _FMT_NUM, unit="m²")
    linha = _kv(ws, linha, "  Sistema viario",         p.areas.area_sistema_viario_m2, _FMT_NUM, unit="m²")
    linha = _kv(ws, linha, "  Area verde / institucional",
                p.areas.area_verde_m2 + p.areas.area_institucional_m2, _FMT_NUM, unit="m²")
    linha = _kv(ws, linha, "  Area APP",               p.areas.area_app_m2,            _FMT_NUM, unit="m²")
    linha = _kv(ws, linha, "Area total de lotes",      p.areas.area_lotes_m2,          _FMT_NUM, unit="m²")
    linha = _kv(ws, linha, "Aproveitamento (%)",       p.areas.aproveitamento * 100,   _FMT_NUM, unit="%")
    linha += 1

    linha = _sec(ws, linha, "TIPOLOGIAS", ncols=3)
    ws.cell(linha, 1, "Tipologia").font = _F_B10
    ws.cell(linha, 2, "Qtd").font = _F_B10
    ws.cell(linha, 2).alignment = Alignment(horizontal="right")
    ws.cell(linha, 3, "Area m² / VGV Lote").font = _F_B10
    for col in (1, 2, 3):
        ws.cell(linha, col).fill = _fill(_C_STONE_D)
    linha += 1
    for tip in p.tipologias:
        ws.cell(linha, 1, tip.nome).font = _F_N10
        ws.cell(linha, 2, tip.quantidade).font = _F_N10
        ws.cell(linha, 2).number_format = _FMT_INT
        ws.cell(linha, 2).alignment = Alignment(horizontal="right")
        ws.cell(linha, 3, f"{tip.area_lote_m2:.0f} m²  |  R$ {int(tip.vgv_lote):,}".replace(",", ".")).font = _F_N10
        linha += 1
    linha = _kv(ws, linha, "TOTAL LOTES", p.total_lotes, _FMT_INT, bold=True)
    linha = _kv(ws, linha, "VGV Bruto total", p.vgv_bruto, _FMT_RS, bold=True)
    linha += 1

    linha = _sec(ws, linha, "DATAS DO PROJETO", ncols=3)
    d = p.datas
    linha = _kv(ws, linha, "Inicio do projeto",   d.inicio_projeto.strftime("%m/%Y"))
    linha = _kv(ws, linha, "Aprovacao",            d.aprovacao.strftime("%m/%Y"))
    linha = _kv(ws, linha, "Lancamento de vendas", d.lancamento_vendas.strftime("%m/%Y"))
    linha = _kv(ws, linha, "Inicio das obras",     d.inicio_obras.strftime("%m/%Y"))
    linha = _kv(ws, linha, "Termino das obras",    d.termino_obras.strftime("%m/%Y"))
    linha += 1

    # ---- AQUISICAO DO TERRENO ----
    aq = projeto.aquisicao
    linha = _sec(ws, linha, "AQUISICAO DO TERRENO", ncols=3)
    linha = _kv(ws, linha, "Valor total",        aq.valor_total,      _FMT_RS,  bold=True)
    linha = _kv(ws, linha, "Forma de pagamento", aq.forma_pagamento)
    if aq.forma_pagamento == "a_vista":
        linha = _kv(ws, linha, "  Mes de pagamento", aq.mes_pagamento, _FMT_INT, unit="mês")
    elif aq.forma_pagamento == "parcelado":
        linha = _kv(ws, linha, "  Mes inicio parcelas", aq.mes_inicio_parcelas, _FMT_INT, unit="mês")
        linha = _kv(ws, linha, "  Qtd parcelas",        aq.qtd_parcelas,        _FMT_INT, unit="parcelas")
    elif aq.forma_pagamento == "customizado" and aq.fluxo_percentuais:
        linha = _kv(ws, linha, "  Mes inicio",          aq.fluxo_mes_inicio,    _FMT_INT, unit="mês")
        linha = _kv(ws, linha, "  Parcelas definidas",  len(aq.fluxo_percentuais), _FMT_INT)
    elif aq.forma_pagamento == "sem_desembolso":
        linha = _kv(ws, linha, "  Sem desembolso de caixa — custo embutido no VGV permutado")
    if aq.custo_cartorio > 0:
        linha = _kv(ws, linha, "Custo cartorio",         aq.custo_cartorio,      _FMT_RS)
        linha = _kv(ws, linha, "  Mes pagamento cartorio", aq.mes_pagamento_cartorio, _FMT_INT, unit="mês")

    # ---- FLUXO (lado direito) — Receitas do projeto ----
    # Rotulado explicitamente para nao ser confundido com desembolso do terreno
    colunas_fluxo = ["Desemb. Terreno", "Rec. Nominal Venda", "Total Entradas"]
    df_cols = {
        "Desemb. Terreno":   "Aquisicao Terreno",
        "Rec. Nominal Venda": "Receita Nominal Venda",
        "Total Entradas":    "Total Entradas",
    }
    meses = df["Mes"].tolist()
    series = [[df[col].iloc[i] for i in range(len(df))] for col in df_cols.values() if col in df.columns]
    if len(series) < len(colunas_fluxo):
        series = series + [[0.0] * len(meses)] * (len(colunas_fluxo) - len(series))

    row_flow = _cabecalho_fluxo(ws, 3, colunas_fluxo)
    _linhas_fluxo(ws, row_flow, meses, series)


# =====================================================================
# 3. RECEITAS
# =====================================================================

def _aba_receitas(wb: Workbook, projeto: Projeto, resultado: ResultadoCalculo) -> None:
    ws = wb.create_sheet("Receitas")
    _dim_auditoria(ws)

    rec = projeto.receitas
    df  = resultado.fluxo_caixa
    linha = _titulo_projeto(ws, projeto)

    linha = _sec(ws, linha, "FLUXOS DE RECEBIVEIS POR TIPOLOGIA", ncols=3)
    for ft in rec.fluxos_tipologia:
        linha = _sub(ws, linha, ft.nome_tipologia, ncols=3)
        linha = _kv(ws, linha, "  Sinal",                ft.percentual_sinal,          _FMT_NUM, unit="%")
        linha = _kv(ws, linha, "    Parcelas de sinal",  ft.qtd_parcelas_sinal,         _FMT_INT, unit="parcelas")
        linha = _kv(ws, linha, "  Parcelas durante obra", ft.percentual_obra,           _FMT_NUM, unit="%")
        linha = _kv(ws, linha, "    Juros parcelas obra", ft.juros_parcelas_obra_am,    _FMT_NUM, unit="% a.m.")
        linha = _kv(ws, linha, "  Baloes anuais",         ft.percentual_baloes,         _FMT_NUM, unit="%")
        linha = _kv(ws, linha, "    Qtd baloes",          ft.qtd_baloes,                _FMT_INT)
        linha = _kv(ws, linha, "  Financiamento pos-obra", ft.percentual_financiamento, _FMT_NUM, unit="%")
        linha = _kv(ws, linha, "    Parcelas financiamento", ft.qtd_parcelas_financiamento, _FMT_INT)
        linha = _kv(ws, linha, "    Juros financiamento", ft.juros_financiamento_am,    _FMT_NUM, unit="% a.m.")
        soma = ft.soma_fluxo
        linha = _kv(ws, linha, "  SOMA (deve ser 100%)", soma, _FMT_NUM, bold=True,
                    cor_val=_C_GREEN if abs(soma - 100) < 0.01 else _C_RED)
        linha = _kv(ws, linha, "  Soma curva de vendas (%)", ft.soma_curva, _FMT_NUM,
                    cor_val=_C_GREEN if abs(ft.soma_curva - 100) < 0.01 else _C_RED)
        linha += 1

    # Permuta
    linha += 1
    linha = _sec(ws, linha, "PERMUTA", ncols=3)
    linha = _kv(ws, linha, "Tipo de permuta", rec.tipo_permuta)
    if rec.tipo_permuta == "fisica" and rec.permuta_fisica:
        for pm in rec.permuta_fisica:
            linha = _kv(ws, linha, f"  {pm.tipologia}", pm.percentual, _FMT_NUM, unit="% dos lotes")

    # ---- FLUXO ----
    colunas_fluxo = ["Rec. Nominal", "Rec. Juros", "Total Entradas"]
    df_cols_vals = []
    for col_name in ["Receita Nominal Venda", "Receita Financeira (Juros)", "Total Entradas"]:
        if col_name in df.columns:
            df_cols_vals.append(df[col_name].tolist())
        else:
            df_cols_vals.append([0.0] * len(df))
    meses = df["Mes"].tolist()
    row_flow = _cabecalho_fluxo(ws, 3, colunas_fluxo)
    _linhas_fluxo(ws, row_flow, meses, df_cols_vals)


# =====================================================================
# 4. OBRAS
# =====================================================================

def _aba_obras(wb: Workbook, projeto: Projeto, resultado: ResultadoCalculo) -> None:
    ws = wb.create_sheet("Obras")
    _dim_auditoria(ws)

    obras = projeto.obras
    df    = resultado.fluxo_caixa
    linha = _titulo_projeto(ws, projeto)

    multiplicador = (
        (1 + obras.bdi_percentual / 100)
        * (1 + obras.contingencia_percentual / 100)
    )

    linha = _sec(ws, linha, "PARAMETROS DE OBRAS", ncols=3)
    linha = _kv(ws, linha, "Modo", "Detalhado" if obras.modo == "detalhado" else "Resumido")
    linha = _kv(ws, linha, "BDI (%)", obras.bdi_percentual, _FMT_NUM, unit="%")
    linha = _kv(ws, linha, "Contingencia (%)", obras.contingencia_percentual, _FMT_NUM, unit="%")
    linha = _kv(ws, linha, "Multiplicador (BDI x Contingencia)", multiplicador, _FMT_NUM, bold=True)
    linha += 1

    horizonte = len(df) - 1
    meses     = df["Mes"].tolist()

    if obras.modo == "detalhado" and obras.etapas:
        linha = _sec(ws, linha, "ETAPAS", ncols=3)
        ws.cell(linha, 1, "Etapa").font = _F_B10
        ws.cell(linha, 2, "Valor Direto (R$)").font = _F_B10
        ws.cell(linha, 2).alignment = Alignment(horizontal="right")
        ws.cell(linha, 3, "M.Ini / Dur / Curva").font = _F_B10
        for col in (1, 2, 3):
            ws.cell(linha, col).fill = _fill(_C_STONE_D)
        linha += 1

        custo_direto_total = 0.0
        for et in obras.etapas:
            ws.cell(linha, 1, et.nome).font = _F_N10
            ws.cell(linha, 2, et.valor_total).number_format = _FMT_RS
            ws.cell(linha, 2).font = _F_N10
            ws.cell(linha, 2).alignment = Alignment(horizontal="right")
            ws.cell(linha, 3, f"M{et.mes_inicio}  /  {et.duracao_meses}m  /  {et.curva}").font = _F_S9
            custo_direto_total += et.valor_total
            linha += 1

        linha = _kv(ws, linha, "Custo Direto Total", custo_direto_total, _FMT_RS, bold=True)
        linha = _kv(ws, linha, "Custo Total (com BDI e Contingencia)",
                    custo_direto_total * multiplicador, _FMT_RS, bold=True)

        # Fluxo por etapa
        colunas_fluxo = [et.nome for et in obras.etapas] + ["TOTAL OBRAS"]
        series = []
        total_vec = np.zeros(horizonte + 1)
        for et in obras.etapas:
            vec = _distribuir_etapa(et, horizonte) * multiplicador
            total_vec += vec
            series.append([float(vec[m]) for m in meses])
        series.append([float(total_vec[m]) for m in meses])

    else:  # resumido
        res = obras.resumido
        if res:
            areas = projeto.terreno.areas
            area_base = (
                areas.area_sistema_viario_m2
                if res.base_calculo == "sistema_viario"
                else areas.area_lotes_m2
            )
            custo_direto_res = area_base * res.valor_por_m2
            custo_total_res  = custo_direto_res * multiplicador

            linha = _sec(ws, linha, "ORCAMENTO RESUMIDO", ncols=3)
            linha = _kv(ws, linha, "Base de calculo",     res.base_calculo)
            linha = _kv(ws, linha, "Area utilizada",      area_base,         _FMT_NUM, unit="m²")
            linha = _kv(ws, linha, "Valor por m2",        res.valor_por_m2,  _FMT_RS2, unit="R$/m²")
            linha = _kv(ws, linha, "Custo direto",        custo_direto_res,  _FMT_RS,  bold=False)
            linha = _kv(ws, linha, "Custo total (c/ BDI e Contingencia)", custo_total_res, _FMT_RS, bold=True)
            linha = _kv(ws, linha, "Mes inicio",          res.mes_inicio,    _FMT_INT, unit="mês")
            linha = _kv(ws, linha, "Duracao",             res.duracao_meses, _FMT_INT, unit="meses")

        colunas_fluxo = ["TOTAL OBRAS"]
        if "Obras" in df.columns:
            series = [df["Obras"].tolist()]
        else:
            series = [[0.0] * len(meses)]

    row_flow = _cabecalho_fluxo(ws, 3, colunas_fluxo, largura=14)
    _linhas_fluxo(ws, row_flow, meses, series)


# =====================================================================
# 5. DESENVOLVIMENTO
# =====================================================================

def _aba_desenvolvimento(wb: Workbook, projeto: Projeto, resultado: ResultadoCalculo) -> None:
    ws = wb.create_sheet("Desenvolvimento")
    _dim_auditoria(ws)

    dev = projeto.desenvolvimento
    df  = resultado.fluxo_caixa
    linha = _titulo_projeto(ws, projeto)

    horizonte = len(df) - 1
    meses     = df["Mes"].tolist()

    linha = _sec(ws, linha, "DESPESAS DE LOTEAMENTO", ncols=3)
    ws.cell(linha, 1, "Despesa").font = _F_B10
    ws.cell(linha, 2, "Valor Total (R$)").font = _F_B10
    ws.cell(linha, 2).alignment = Alignment(horizontal="right")
    ws.cell(linha, 3, "Categoria / Modo").font = _F_B10
    for col in (1, 2, 3):
        ws.cell(linha, col).fill = _fill(_C_STONE_D)
    linha += 1

    total_desp = 0.0
    for desp in dev.despesas:
        ws.cell(linha, 1, desp.nome).font = _F_N10
        ws.cell(linha, 2, desp.valor_total).number_format = _FMT_RS
        ws.cell(linha, 2).font = _F_N10
        ws.cell(linha, 2).alignment = Alignment(horizontal="right")
        ws.cell(linha, 3, f"{desp.categoria}  /  {desp.modo}").font = _F_S9
        total_desp += desp.valor_total
        linha += 1

    linha = _kv(ws, linha, "Total Despesas", total_desp, _FMT_RS, bold=True)
    linha += 1

    linha = _sec(ws, linha, "ADMINISTRACAO", ncols=3)
    linha = _kv(ws, linha, "Percentual sobre receita mensal",
                dev.administracao.percentual, _FMT_NUM, unit="%")
    linha += 1

    # Fluxo: categorias do fluxo_caixa + admin
    cat_map = {
        "Projetos":              "Projetos",
        "Licenciamento":         "Licenciamento",
        "Marketing":             "Marketing",
        "Outros Desenv.":        "Outros Desenvolvimento",
        "Administracao":         "Administracao",
    }
    colunas_fluxo = list(cat_map.keys()) + ["TOTAL DESENV."]
    series = []
    total_vec = [0.0] * len(meses)
    for label, col_df in cat_map.items():
        if col_df in df.columns:
            vals = df[col_df].tolist()
        else:
            vals = [0.0] * len(meses)
        series.append(vals)
        total_vec = [total_vec[i] + vals[i] for i in range(len(meses))]
    series.append(total_vec)

    row_flow = _cabecalho_fluxo(ws, 3, colunas_fluxo, largura=14)
    _linhas_fluxo(ws, row_flow, meses, series)


# =====================================================================
# 6. IMPOSTOS
# =====================================================================

def _aba_impostos(wb: Workbook, projeto: Projeto, resultado: ResultadoCalculo) -> None:
    ws = wb.create_sheet("Impostos")
    _dim_auditoria(ws)

    imp = projeto.impostos
    df  = resultado.fluxo_caixa
    linha = _titulo_projeto(ws, projeto)

    linha = _sec(ws, linha, "REGIME TRIBUTARIO", ncols=3)
    linha = _kv(ws, linha, "Regime", imp.tributos.regime)
    linha = _kv(ws, linha, "Aliquota efetiva (%)", imp.tributos.aliquota_efetiva, _FMT_NUM, unit="%")
    linha = _kv(ws, linha, "Regime de apuracao", imp.tributos.regime_apuracao)
    linha += 1

    linha = _sec(ws, linha, "COMISSAO DE VENDAS", ncols=3)
    com = imp.comissao
    linha = _kv(ws, linha, "Percentual sobre VGV (%)", com.percentual_vgv, _FMT_NUM, unit="%")
    linha = _kv(ws, linha, "Modo de pagamento", com.modo_pagamento)
    if com.modo_pagamento == "misto":
        linha = _kv(ws, linha, "  % no ato", com.misto_percentual_no_ato, _FMT_NUM, unit="%")
        linha = _kv(ws, linha, "  Parcelas diluidas", com.misto_qtd_parcelas, _FMT_INT)
    linha += 1

    if imp.permuta_financeira:
        pf = imp.permuta_financeira
        linha = _sec(ws, linha, "PERMUTA FINANCEIRA", ncols=3)
        linha = _kv(ws, linha, "Percentual do VGV (%)", pf.percentual_vgv, _FMT_NUM, unit="%")
        linha = _kv(ws, linha, "Modo de pagamento", pf.modo_pagamento)

    # Fluxo
    col_map = {
        "Comissao":         "Comissao",
        "Impostos":         "Impostos",
        "Permuta Fin.":     "Permuta Financeira",
    }
    colunas_fluxo = list(col_map.keys()) + ["TOTAL"]
    series = []
    total_vec = [0.0] * len(df)
    meses = df["Mes"].tolist()
    for label, col_df in col_map.items():
        if col_df in df.columns:
            vals = df[col_df].tolist()
        else:
            vals = [0.0] * len(df)
        series.append(vals)
        total_vec = [total_vec[i] + vals[i] for i in range(len(df))]
    series.append(total_vec)

    row_flow = _cabecalho_fluxo(ws, 3, colunas_fluxo, largura=16)
    _linhas_fluxo(ws, row_flow, meses, series)


# =====================================================================
# 7. FLUXO DE CAIXA COMPLETO
# =====================================================================

def _aba_fluxo_caixa(wb: Workbook, projeto: "Projeto", resultado: ResultadoCalculo) -> None:
    ws = wb.create_sheet("Fluxo de Caixa")
    df = resultado.fluxo_caixa

    _linhas_negrito  = {"Total Entradas", "Total Saidas", "Saldo Acumulado"}
    _linhas_verde    = {"Total Entradas"}
    _linhas_vermelha = {"Total Saidas"}
    _linhas_azul     = {"Saldo Acumulado", "Saldo Descontado Acumulado"}
    _sem_totais      = {"Saldo Acumulado", "Saldo Descontado Acumulado"}

    # Nomes para exibicao com acento
    _NOMES_PT: dict[str, str] = {
        "Receita Nominal Venda":           "Receita Nominal Venda",
        "Receita Financeira (Juros)":      "Receita Financeira (Juros)",
        "Correcao Monetaria (Parcelas)":   "Correção Monetária (Parcelas)",
        "Outras Receitas":                 "Outras Receitas",
        "Total Entradas":                  "Total Entradas",
        "Aquisicao Terreno":               "Aquisição Terreno",
        "Cartorio":                        "Cartório",
        "Obras":                           "Obras",
        "Projetos":                        "Projetos",
        "Licenciamento":                   "Licenciamento",
        "Marketing":                       "Marketing",
        "Outros Desenvolvimento":          "Outros Desenvolvimento",
        "Administracao":                   "Administração",
        "Comissao":                        "Comissão",
        "Impostos":                        "Impostos",
        "Permuta Financeira":              "Permuta Financeira",
        "Saque Financiamento":             "Saque Financiamento",
        "Amortizacao Financiamento":       "Amortização Financiamento",
        "Juros Financiamento Banco":       "Juros Financiamento (Banco)",
        "Comissao Abertura Financiamento": "Comissão Abertura Financiamento",
        "Saque Investidor":                "Saque Investidor",
        "Amortizacao Investidor":          "Amortização Investidor",
        "Juros Investidor":                "Juros Investidor",
        "Total Saidas":                    "Total Saídas",
        "Saldo do Mes":                    "Saldo do Mês",
        "Saldo Acumulado":                 "Saldo Acumulado",
        "Saldo Descontado Acumulado":      "Saldo Descontado Acumulado",
    }

    # Ordem das linhas de entrada e saida
    _ENTRADAS_COLS = [
        "Receita Nominal Venda", "Receita Financeira (Juros)",
        "Correcao Monetaria (Parcelas)", "Outras Receitas",
        "Saque Financiamento", "Saque Investidor",
    ]
    _SAIDAS_COLS = [
        "Aquisicao Terreno", "Cartorio", "Obras", "Projetos", "Licenciamento",
        "Marketing", "Outros Desenvolvimento", "Administracao",
        "Amortizacao Financiamento", "Juros Financiamento Banco",
        "Comissao Abertura Financiamento", "Amortizacao Investidor", "Juros Investidor",
        "Comissao", "Impostos", "Permuta Financeira",
    ]

    meses_lista = df["Mes"].tolist()
    n_meses = len(meses_lista)

    vgv_nominal = float(df["Receita Nominal Venda"].sum()) if "Receita Nominal Venda" in df.columns else 1.0
    row_totals = {c: float(df[c].sum()) for c in df.columns if c != "Mes"}

    def _show_row(col_name: str) -> bool:
        if col_name in {"Total Entradas", "Total Saidas", "Saldo do Mes",
                        "Saldo Acumulado", "Saldo Descontado Acumulado"}:
            return True
        return abs(row_totals.get(col_name, 0.0)) > 0.5

    # Marcos
    try:
        inicio = projeto.terreno.datas.inicio_projeto
        _marcos_dict: dict[int, str] = {
            0: "Início",
            meses_entre(inicio, projeto.terreno.datas.aprovacao):        "Aprovação",
            meses_entre(inicio, projeto.terreno.datas.lancamento_vendas): "Lançamento",
            meses_entre(inicio, projeto.terreno.datas.inicio_obras):     "Ini. Obras",
            meses_entre(inicio, projeto.terreno.datas.termino_obras):    "Fim Obras",
        }
    except Exception:
        _marcos_dict = {}

    # Labels de calendario (set/26, out/26, ...)
    _MES_NOMES = ["jan","fev","mar","abr","mai","jun","jul","ago","set","out","nov","dez"]
    try:
        inicio = projeto.terreno.datas.inicio_projeto
        _cal_labels: list[str] = []
        for _mes_int in meses_lista:
            _m = int(_mes_int)
            _total_m = inicio.month - 1 + _m
            _y = inicio.year + _total_m // 12
            _mo = _total_m % 12 + 1
            _cal_labels.append(f"{_MES_NOMES[_mo - 1]}/{str(_y)[2:]}")
    except Exception:
        _cal_labels = [f"M{int(m)}" for m in meses_lista]

    # Layout: col A=Item, B=% Nominal, C=Total, D+=meses
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 16
    for _j in range(n_meses):
        ws.column_dimensions[get_column_letter(4 + _j)].width = 11

    # --- Linha 1: indices de mes (M0, M1, ...) ---
    ws.row_dimensions[1].height = 18
    for _col_num, _label in [(1, "Item"), (2, "% Nominal"), (3, "Total")]:
        _c = ws.cell(1, _col_num, _label)
        _c.font = _F_W10
        _c.fill = _fill(_C_DARK)
        _c.alignment = Alignment(
            horizontal="left" if _col_num == 1 else "center",
            vertical="center",
            indent=1 if _col_num == 1 else 0,
        )
    for _j, _mes in enumerate(meses_lista, start=4):
        _c = ws.cell(1, _j, f"M{int(_mes)}")
        _c.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        _c.fill = _fill(_C_DARK)
        _c.alignment = Alignment(horizontal="center", vertical="center")

    # --- Linha 2: datas do calendario ---
    ws.row_dimensions[2].height = 15
    for _col_num in range(1, 4):
        ws.cell(2, _col_num).fill = _fill(_C_DARK)
    for _j, _cal in enumerate(_cal_labels, start=4):
        _c = ws.cell(2, _j, _cal)
        _c.font = Font(name="Calibri", size=9, color="BBBBBB")
        _c.fill = _fill(_C_DARK)
        _c.alignment = Alignment(horizontal="center", vertical="center")

    # --- Linha 3: marcos do projeto ---
    _C_OCHRE_F = "F5EED8"
    ws.row_dimensions[3].height = 14
    _c = ws.cell(3, 1, "Marcos")
    _c.font = Font(name="Calibri", size=9, italic=True, color=_C_OCHRE)
    _c.fill = _fill(_C_OCHRE_F)
    _c.alignment = Alignment(indent=1)
    for _col_num in [2, 3]:
        ws.cell(3, _col_num).fill = _fill(_C_OCHRE_F)
    for _j, _mes in enumerate(meses_lista, start=4):
        _label = _marcos_dict.get(int(_mes), "")
        _c = ws.cell(3, _j, _label or None)
        _c.fill = _fill(_C_OCHRE_F)
        if _label:
            _c.font = Font(name="Calibri", size=9, bold=True, color=_C_OCHRE)
            _c.alignment = Alignment(horizontal="center")

    # --- Helper: escrever uma linha de dados ---
    def _escrever_linha(linha: int, col_name: str, alternado: bool = False) -> None:
        negrito = col_name in _linhas_negrito
        sem_tot = col_name in _sem_totais
        cor = (
            _C_GREEN   if col_name in _linhas_verde
            else _C_RED if col_name in _linhas_vermelha
            else _C_STONE if col_name in _linhas_azul
            else (_C_STONE_D if alternado else _C_WHITE)
        )
        ws.row_dimensions[linha].height = 15

        _ca = ws.cell(linha, 1, _NOMES_PT.get(col_name, col_name))
        _ca.font = _F_B10 if negrito else _F_N10
        _ca.fill = _fill(cor)
        _ca.alignment = Alignment(indent=1, vertical="center")

        total = row_totals.get(col_name, 0.0)
        pct_val = (total / vgv_nominal) if (not sem_tot and vgv_nominal) else None

        _cb = ws.cell(linha, 2, round(pct_val, 4) if pct_val is not None and abs(pct_val) > 0.0001 else None)
        _cb.number_format = _FMT_PCT
        _cb.font = _F_B10 if negrito else _F_N10
        _cb.fill = _fill(cor)
        _cb.alignment = Alignment(horizontal="right", vertical="center")

        _cc = ws.cell(linha, 3, total if (not sem_tot and abs(total) > 0.5) else None)
        _cc.number_format = _FMT_RS
        _cc.font = _F_B10 if negrito else _F_N10
        _cc.fill = _fill(cor)
        _cc.alignment = Alignment(horizontal="right", vertical="center")

        for _j, _val in enumerate(df[col_name].tolist(), start=4):
            _v = float(_val)
            _cv = ws.cell(linha, _j, _v if abs(_v) > 0.5 else None)
            _cv.number_format = _FMT_RS
            _cv.font = _F_B10 if negrito else _F_N10
            _cv.fill = _fill(cor)
            _cv.alignment = Alignment(horizontal="right", vertical="center")

    def _escrever_separador(linha: int) -> None:
        ws.row_dimensions[linha].height = 5
        for _col_num in range(1, 4 + n_meses):
            ws.cell(linha, _col_num).fill = _fill("F0EEE9")

    # --- Dados ---
    cur = 4

    alt = False
    for col_name in _ENTRADAS_COLS:
        if col_name in df.columns and _show_row(col_name):
            _escrever_linha(cur, col_name, alternado=alt)
            alt = not alt
            cur += 1
    _escrever_linha(cur, "Total Entradas")
    cur += 1
    _escrever_separador(cur); cur += 1

    alt = False
    for col_name in _SAIDAS_COLS:
        if col_name in df.columns and _show_row(col_name):
            _escrever_linha(cur, col_name, alternado=alt)
            alt = not alt
            cur += 1
    _escrever_linha(cur, "Total Saidas")
    cur += 1
    _escrever_separador(cur); cur += 1

    _escrever_linha(cur, "Saldo do Mes"); cur += 1
    _escrever_linha(cur, "Saldo Acumulado"); cur += 1
    _escrever_separador(cur); cur += 1
    _escrever_linha(cur, "Saldo Descontado Acumulado")

    ws.freeze_panes = "D4"


# =====================================================================
# 8. VERIFICACAO DE RECEITAS
# =====================================================================

def _detalhar_venda(
    vgv_venda: float, mes_venda: int, fluxo, mes_termino_obras: int
) -> list[dict]:
    eventos: list[dict] = []

    valor_sinal = vgv_venda * (fluxo.percentual_sinal / 100)
    if valor_sinal > 0 and fluxo.qtd_parcelas_sinal > 0:
        parcela_s = valor_sinal / fluxo.qtd_parcelas_sinal
        for k in range(fluxo.qtd_parcelas_sinal):
            eventos.append({"tipo": "Sinal", "mes": mes_venda + k,
                            "valor": parcela_s, "principal": parcela_s,
                            "juros": 0.0, "saldo_devedor": 0.0})

    valor_obra = vgv_venda * (fluxo.percentual_obra / 100)
    if valor_obra > 0:
        mes_ini_obra = mes_venda + max(fluxo.qtd_parcelas_sinal, 1)
        qtd_obra = mes_termino_obras - mes_ini_obra + 1
        if qtd_obra > 0:
            i = fluxo.juros_parcelas_obra_am / 100
            parc = parcela_price(valor_obra, i, qtd_obra)
            saldo = valor_obra
            for k in range(qtd_obra):
                j = saldo * i
                amort = parc - j
                saldo = max(saldo - amort, 0.0)
                eventos.append({"tipo": "Parcela Obra", "mes": mes_ini_obra + k,
                                "valor": parc, "principal": amort,
                                "juros": j, "saldo_devedor": saldo})
        else:
            eventos.append({"tipo": "Parcela Obra (concentrada)", "mes": mes_ini_obra,
                            "valor": valor_obra, "principal": valor_obra,
                            "juros": 0.0, "saldo_devedor": 0.0})

    valor_baloes = vgv_venda * (fluxo.percentual_baloes / 100)
    if valor_baloes > 0 and fluxo.qtd_baloes > 0:
        valor_balao = valor_baloes / fluxo.qtd_baloes
        for k in range(1, fluxo.qtd_baloes + 1):
            eventos.append({"tipo": f"Balao #{k}", "mes": mes_venda + k * 12,
                            "valor": valor_balao, "principal": valor_balao,
                            "juros": 0.0, "saldo_devedor": 0.0})

    valor_fin = vgv_venda * (fluxo.percentual_financiamento / 100)
    if valor_fin > 0 and fluxo.qtd_parcelas_financiamento > 0:
        i = fluxo.juros_financiamento_am / 100
        parc_fin = parcela_price(valor_fin, i, fluxo.qtd_parcelas_financiamento)
        saldo = valor_fin
        for k in range(fluxo.qtd_parcelas_financiamento):
            j = saldo * i
            amort = parc_fin - j
            saldo = max(saldo - amort, 0.0)
            eventos.append({"tipo": "Financiamento", "mes": mes_termino_obras + 1 + k,
                            "valor": parc_fin, "principal": amort,
                            "juros": j, "saldo_devedor": saldo})

    eventos.sort(key=lambda x: x["mes"])
    return eventos


def _aba_verificacao_receitas(
    wb: Workbook, projeto: Projeto, resultado: ResultadoCalculo
) -> None:
    ws = wb.create_sheet("Verificacao Receitas")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20

    terreno  = projeto.terreno
    receitas = projeto.receitas
    r        = resultado.resumo

    vgv_vendavel = calcular_vgv_vendavel(terreno, receitas)
    mes_termino  = meses_entre(terreno.datas.inicio_projeto, terreno.datas.termino_obras)

    mapa_permuta: dict[str, float] = {}
    if receitas.tipo_permuta == "fisica":
        mapa_permuta = {p.tipologia: p.percentual for p in receitas.permuta_fisica}
    mapa_tipologias = {t.nome: t for t in terreno.tipologias}
    total_lotes_vendaveis = sum(
        tip.quantidade * (1 - mapa_permuta.get(tip.nome, 0.0) / 100)
        for tip in terreno.tipologias
    )

    # Novo formato: usa fluxos_tipologia; legado: usa fluxos_recebiveis + curva_vendas
    usa_novo_formato = bool(receitas.fluxos_tipologia)
    mapa_fluxos = {ft.nome_tipologia: ft.as_fluxo_recebiveis() for ft in receitas.fluxos_tipologia} if usa_novo_formato else {f.nome: f for f in receitas.fluxos_recebiveis}

    _hdr(ws, 1, 1, "VERIFICACAO DE RECEITAS — FLUXO DETALHADO POR MES DE VENDA",
         _C_DARK, fonte=_F_W11, ncols=6, height=24)

    def _rs(v: float) -> str:
        n = int(round(abs(v)))
        s = f"{n:,}".replace(",", ".")
        return f"R$ {s}" if v >= 0 else f"-R$ {s}"

    vgv_efetivo = r.get("vgv_efetivo_vendavel", vgv_vendavel)
    info = (
        f"VGV Vendavel (base): {_rs(vgv_vendavel)}"
        + (f"    VGV Efetivo: {_rs(vgv_efetivo)}" if abs(vgv_efetivo - vgv_vendavel) > 1 else "")
        + f"    Principal: {_rs(r['receita_nominal_venda'])}"
        + f"    Juros: {_rs(r['receita_financeira'])}"
        + f"    Total: {_rs(r['vgv_total_recebido'])}"
    )
    cel_info = ws.cell(2, 1, info)
    cel_info.font = _F_B10
    ws.merge_cells("A2:F2")

    linha = 4
    total_principal = 0.0
    total_juros     = 0.0
    total_valor     = 0.0

    def _escrever_venda(fluxo, mes_venda: int, vgv_por_mes: float, lotes_por_mes: float) -> None:
        nonlocal linha, total_principal, total_juros, total_valor
        eventos = _detalhar_venda(vgv_por_mes, mes_venda, fluxo, mes_termino)
        if not eventos:
            return

        sub_valor     = sum(e["valor"]     for e in eventos)
        sub_principal = sum(e["principal"] for e in eventos)
        sub_juros     = sum(e["juros"]     for e in eventos)

        partes = [f"M{mes_venda}", f"Fluxo: {fluxo.nome}",
                  f"VGV: {_rs(vgv_por_mes)}", f"Lotes: {lotes_por_mes:.3f}"]
        if sub_juros > 0.01:
            partes.append(f"Juros: {_rs(sub_juros)}")
        cel = ws.cell(linha, 1, "   ".join(partes))
        cel.font = _F_W10
        cel.fill = _fill(_C_OCHRE)
        ws.merge_cells(f"A{linha}:F{linha}")
        ws.row_dimensions[linha].height = 18
        linha += 1

        for ci, txt in enumerate(["Tipo", "Mes Receb.", "Valor (R$)",
                                   "Principal (R$)", "Juros (R$)", "Saldo Dev. (R$)"], start=1):
            c = ws.cell(linha, ci, txt)
            c.font = _F_B10
            c.fill = _fill(_C_STONE_D)
        linha += 1

        for ev in eventos:
            ws.cell(linha, 1, ev["tipo"]).font = _F_N10
            ws.cell(linha, 2, ev["mes"]).number_format = _FMT_MES
            ws.cell(linha, 2).alignment = Alignment(horizontal="center")
            for ci, key in [(3, "valor"), (4, "principal"), (5, "juros"), (6, "saldo_devedor")]:
                v = ev[key]
                c = ws.cell(linha, ci, v if v != 0.0 else None)
                c.number_format = _FMT_RS2
                if key == "juros" and v > 0.01:
                    c.fill = _fill("FFF0CC")
            linha += 1

        ws.cell(linha, 1, "Subtotal").font = _F_B10
        ws.cell(linha, 1).fill = _fill(_C_STONE_D)
        for ci, v in [(3, sub_valor), (4, sub_principal), (5, sub_juros)]:
            c = ws.cell(linha, ci, v if v != 0.0 else None)
            c.font = _F_B10
            c.number_format = _FMT_RS2
            c.fill = _fill(_C_GREEN if ci == 4 else _C_STONE_D)
        linha += 2

        total_principal += sub_principal
        total_juros     += sub_juros
        total_valor     += sub_valor

    if usa_novo_formato:
        for ft in receitas.fluxos_tipologia:
            fluxo = mapa_fluxos.get(ft.nome_tipologia)
            if fluxo is None:
                continue
            tip = mapa_tipologias.get(ft.nome_tipologia)
            if tip is None:
                continue
            pct_perm = mapa_permuta.get(ft.nome_tipologia, 0.0)
            vgv_tip  = tip.vgv_total * (1 - pct_perm / 100)
            lotes_tip = tip.quantidade * (1 - pct_perm / 100)
            for mes_str, pct in ft.curva_mensal.items():
                if pct <= 0:
                    continue
                mes_venda = int(mes_str)
                fator = float(ft.fatores_preco.get(mes_venda, ft.fatores_preco.get(str(mes_venda), 1.0))) if ft.fatores_preco else 1.0
                _escrever_venda(fluxo, mes_venda, vgv_tip * (pct / 100) * fator, lotes_tip * (pct / 100))
    else:
        for faixa in receitas.curva_vendas:
            if faixa.fluxo_recebiveis not in mapa_fluxos:
                continue
            fluxo = mapa_fluxos[faixa.fluxo_recebiveis]
            qtd_meses = max(1, faixa.mes_fim - faixa.mes_inicio + 1)
            vgv_por_mes   = vgv_vendavel * (faixa.percentual_estoque / 100) * faixa.fator_preco / qtd_meses
            lotes_por_mes = total_lotes_vendaveis * (faixa.percentual_estoque / 100) / qtd_meses
            for mes_venda in range(faixa.mes_inicio, faixa.mes_fim + 1):
                _escrever_venda(fluxo, mes_venda, vgv_por_mes, lotes_por_mes)

    # Total geral
    _hdr(ws, linha, 1, "TOTAL GERAL", _C_DARK, fonte=_F_W10, ncols=6, height=22)
    linha += 1
    for ci, txt in [(3, "Recebivel Total"), (4, "Principal Total"), (5, "Juros Total")]:
        ws.cell(linha, ci, txt).font = _F_B10
    linha += 1
    for ci, v in [(3, total_valor), (4, total_principal), (5, total_juros)]:
        c = ws.cell(linha, ci, v if v != 0.0 else None)
        c.font = Font(name="Calibri", size=11, bold=True, color="1A1916")
        c.number_format = _FMT_RS
        c.fill = _fill(_C_GREEN)
    linha += 2

    # Invariantes
    dif_p = abs(total_principal - r["receita_nominal_venda"])
    dif_j = abs(total_juros     - r["receita_financeira"])
    ok_p  = "OK" if dif_p < 1.0 else f"DIVERGENCIA: {dif_p:,.0f}"
    ok_j  = "OK" if dif_j < 1.0 else f"DIVERGENCIA: {dif_j:,.0f}"
    ws.cell(linha,     1, f"Verificacao Principal = Receita Nominal: {ok_p}").font = _F_B10
    ws.cell(linha + 1, 1, f"Verificacao Juros = Rec. Financeira:  {ok_j}").font = _F_B10

    ws.freeze_panes = "A4"


# =====================================================================
# 9. SIMULACAO DE 1 LOTE
# =====================================================================

def _aba_simulacao_lote(wb: Workbook, projeto: Projeto) -> None:
    ws = wb.create_sheet("Simulacao Lote")
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    _hdr(ws, 1, 1, "SIMULACAO DE 1 LOTE POR TIPOLOGIA", _C_DARK,
         fonte=_F_W11, ncols=6, height=24)

    rec = projeto.receitas
    if rec.fluxos_tipologia:
        ft0 = rec.fluxos_tipologia[0]
        fluxo = ft0.as_fluxo_recebiveis()
        curva = ft0.curva_mensal
        primeiro_mes = int(next(iter(curva))) if curva else 0
        mes_venda = primeiro_mes
    elif rec.fluxos_recebiveis:
        fluxo = rec.fluxos_recebiveis[0]
        if rec.curva_vendas:
            primeira = rec.curva_vendas[0]
            mes_venda = (primeira.mes_inicio + primeira.mes_fim) // 2
        else:
            mes_venda = 0
    else:
        return

    mes_termino = meses_entre(
        projeto.terreno.datas.inicio_projeto,
        projeto.terreno.datas.termino_obras,
    )

    linha = 3
    ws.cell(linha, 1,
            f"Fluxo: {fluxo.nome}  |  Mes de venda: M{mes_venda}  |  Termino obras: M{mes_termino}"
            ).font = _F_B10
    linha += 2

    for tip in projeto.terreno.tipologias:
        _hdr(ws, linha, 1, f"Tipologia: {tip.nome}  —  VGV/lote: R$ {int(tip.vgv_lote):,}".replace(",", "."),
             _C_OCHRE, fonte=_F_W10, ncols=6, height=18)
        linha += 1

        for ci, txt in enumerate(["Mes", "Tipo", "Parcela (R$)", "Principal (R$)",
                                   "Juros (R$)", "Saldo Devedor (R$)"], start=1):
            c = ws.cell(linha, ci, txt)
            c.font = _F_B10
            c.fill = _fill(_C_STONE_D)
        linha += 1

        cron = simular_lote_unitario(tip.vgv_lote, fluxo, mes_venda, mes_termino)
        for ev in cron:
            ws.cell(linha, 1, ev["mes"]).number_format = _FMT_MES
            ws.cell(linha, 2, ev["tipo"]).font = _F_N10
            for ci, key in [(3, "valor"), (4, "principal"), (5, "juros"), (6, "saldo_devedor")]:
                v = ev[key]
                c = ws.cell(linha, ci, v if v != 0.0 else None)
                c.number_format = _FMT_RS2
                c.font = _F_N10
                c.alignment = Alignment(horizontal="right")
            linha += 1

        # Totais
        tp = sum(e["valor"]     for e in cron)
        tpr = sum(e["principal"] for e in cron)
        tj = sum(e["juros"]     for e in cron)
        ws.cell(linha, 2, "TOTAL").font = _F_B10
        for ci, v in [(3, tp), (4, tpr), (5, tj)]:
            c = ws.cell(linha, ci, v if v != 0.0 else None)
            c.font = _F_B10
            c.number_format = _FMT_RS2
            c.fill = _fill(_C_GREEN)
            c.alignment = Alignment(horizontal="right")
        linha += 2


# =====================================================================
# ENTRY POINT
# =====================================================================

def exportar_para_excel(
    projeto: Projeto, resultado: ResultadoCalculo, caminho_saida: str | Path
) -> None:
    """Cria o arquivo Excel de auditoria completa."""
    caminho_saida = Path(caminho_saida)
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    _dashboard(wb, projeto, resultado)
    _aba_terreno(wb, projeto, resultado)
    _aba_receitas(wb, projeto, resultado)
    _aba_obras(wb, projeto, resultado)
    _aba_desenvolvimento(wb, projeto, resultado)
    _aba_impostos(wb, projeto, resultado)
    _aba_fluxo_caixa(wb, projeto, resultado)
    _aba_verificacao_receitas(wb, projeto, resultado)
    _aba_simulacao_lote(wb, projeto)

    wb.save(caminho_saida)
